# Mainframe_RTv250829_mplfinance.py

import pandas as pd
import numpy as np
import os
import glob
import subprocess
import plotly.graph_objects as go
from plotly.subplots import make_subplots

try:
    import openpyxl
except ImportError:
    openpyxl = None
    
from CBullDivg_Analysis_vectorized import CBullDivg_analysis
from CBullDivg_x2_analysis_vectorized import CBullDivg_x2_analysis
from HBearDivg_analysis_vectorized import HBearDivg_analysis
from HBullDivg_analysis_vectorized import HBullDivg_analysis
from Initialize_RSI_EMA_MACD import Initialize_RSI_EMA_MACD
from Local_Maximas_Minimas import Local_Max_Min

# --- KEINE ÄNDERUNGEN IN DEN FUNKTIONEN ZUR PARAMETERABFRAGE ---
def get_input_file():
    try:
        ps_command = '''
Add-Type -AssemblyName System.Windows.Forms
$openFileDialog = New-Object System.Windows.Forms.OpenFileDialog
$openFileDialog.Filter = "CSV/Parquet files (*.csv;*.parquet)|*.csv;*.parquet|CSV files (*.csv)|*.csv|Parquet files (*.parquet)|*.parquet|All files (*.*)|*.*"
$openFileDialog.Title = "Select CSV input file"
$openFileDialog.InitialDirectory = "C:\\Projekte\\crt_250816\\data\\raw"
$result = $openFileDialog.ShowDialog()
if ($result -eq [System.Windows.Forms.DialogResult]::OK) {
    Write-Output $openFileDialog.FileName
}
'''
        result = subprocess.run(["powershell", "-Command", ps_command], capture_output=True, text=True, timeout=60)
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout.strip()
        else:
            return get_input_file_console()
    except Exception:
        return get_input_file_console()

def get_input_file_console():
    default_file = r"C:\Projekte\crt_250816\data\raw\btc_1day_candlesticks_all.csv"
    print(f"\nDefault file: {default_file}")
    print("Press Enter to use default, or choose from available files:")
    
    csv_files = []
    search_paths = ["*.csv", "*.parquet", "../*/*.csv", "../*/*.parquet", "../../data/raw/*.csv", "../../data/raw/*.parquet", "../../data/*.csv", "../../data/*.parquet"]
    for pattern in search_paths:
        csv_files.extend(glob.glob(pattern))
    
    for i, file in enumerate(csv_files, 1):
        print(f"{i}: {file}")
    print(f"{len(csv_files) + 1}: Enter custom path")
    print(f"0 or Enter: Use default ({default_file})")
    
    while True:
        try:
            choice = input(f"\nSelect file (0-{len(csv_files) + 1}): ").strip()
            if choice == "" or choice == "0":
                return default_file
            choice_num = int(choice)
            if 1 <= choice_num <= len(csv_files):
                return csv_files[choice_num - 1]
            elif choice_num == len(csv_files) + 1:
                custom_path = input("Enter full path to CSV file (or Enter for default): ")
                return custom_path if custom_path.strip() else default_file
            else:
                print("Invalid choice. Please try again.")
        except (ValueError, EOFError):
            return default_file

def get_analysis_parameters():
    import sys
    if len(sys.argv) > 2:
        try:
            candle_percent = float(sys.argv[2])
            macd_percent = float(sys.argv[3]) if len(sys.argv) > 3 else 3.25
            return candle_percent, macd_percent, None
        except (ValueError, IndexError):
            pass
    
    try:
        print("\nEnter analysis parameters (or press Enter for defaults):")
        candle_input = input("Enter Candle tolerance % (default: 0.1): ").strip()
        candle_percent = 0.1 if candle_input == "" else float(candle_input)
        macd_input = input("Enter MACD tolerance % (default: 3.25): ").strip()
        macd_percent = 3.25 if macd_input == "" else float(macd_input)
        
        print("\nOptional second variant for comparison:")
        candle_input2 = input("Enter second Candle tolerance % (leave empty for none): ").strip()
        macd_input2 = input("Enter second MACD tolerance % (leave empty for none): ").strip()
        
        variant2 = None
        if candle_input2 or macd_input2:
            try:
                candle_percent2 = float(candle_input2) if candle_input2 else candle_percent
                macd_percent2 = float(macd_input2) if macd_input2 else macd_percent
                variant2 = (candle_percent2, macd_percent2)
            except ValueError as e:
                print(f"Invalid input for second variant: {e}")
                variant2 = None
        
        return candle_percent, macd_percent, variant2
    except EOFError:
        return 0.1, 3.25, None

def get_analysis_type():
    import sys
    if len(sys.argv) > 1:
        choice = sys.argv[1].lower().strip()
        if choice in ['a', 'b', 'c', 'd', 'e', 'f']:
            return choice
    
    try:
        print("\nSelect analysis type:")
        print("a: CBullDivg_Analysis (Classic Bullish Divergence)")
        print("b: BullDivg_x2_analysis (Extended Bullish Divergence)")
        print("c: HBearDivg_analysis (Hidden Bearish Divergence)")
        print("d: HBullDivg_analysis (Hidden Bullish Divergence)")
        print("e: All analyses (a-d)")
        print("f: DOE (Design of Experiments)")
        
        while True:
            choice = input("\nEnter your choice (a-f): ").lower().strip()
            if choice in ['a', 'b', 'c', 'd', 'e', 'f']:
                return choice
            print("Invalid choice. Please enter a, b, c, d, e, or f.")
    except EOFError:
        return 'e'

def run_analysis(df, analysis_type, window=5, candle_tol=0.1, macd_tol=3.25):
    results = {}
    if analysis_type in ['a', 'e', 'f']:
        CBullDivg_analysis(df, window, candle_tol, macd_tol)
        results['CBullDivg'] = True
    if analysis_type in ['b', 'e', 'f']:
        CBullDivg_x2_analysis(df, window, candle_tol, macd_tol)
        results['CBullDivg_x2'] = True
    if analysis_type in ['c', 'e', 'f']:
        HBearDivg_analysis(df, window, candle_tol, macd_tol)
        results['HBearDivg'] = True
    if analysis_type in ['d', 'e', 'f']:
        HBullDivg_analysis(df, window, candle_tol, macd_tol)
        results['HBullDivg'] = True
    return results

def export_markers_to_csv(df, filename, analysis_results, candle_percent, macd_percent):
    markers = []
    counts = {analysis: {'classic': 0, 'hidden': 0, 'total': 0} for analysis in analysis_results if analysis_results[analysis]}
    
    for i in range(len(df)):
        current_date = df['date'].iloc[i]
        if 'CBullDivg' in analysis_results:
            if "CBullD_gen" in df.columns and pd.notna(df["CBullD_gen"].iloc[i]) and df["CBullD_gen"].iloc[i] == 1:
                markers.append({'Type': 'CBullDivg_Classic', 'Date': current_date, 'Candle_Percent': candle_percent, 'MACD_Percent': macd_percent})
                counts['CBullDivg']['classic'] += 1
            if "CBullD_neg_MACD" in df.columns and pd.notna(df["CBullD_neg_MACD"].iloc[i]) and df["CBullD_neg_MACD"].iloc[i] == 1:
                markers.append({'Type': 'CBullDivg_Hidden', 'Date': current_date, 'Candle_Percent': candle_percent, 'MACD_Percent': macd_percent})
                counts['CBullDivg']['hidden'] += 1
        
        if 'CBullDivg_x2' in analysis_results:
            if "CBullD_x2_gen" in df.columns and pd.notna(df["CBullD_x2_gen"].iloc[i]) and df["CBullD_x2_gen"].iloc[i] == 1:
                markers.append({'Type': 'CBullDivg_x2_Classic', 'Date': current_date, 'Candle_Percent': candle_percent, 'MACD_Percent': macd_percent})
                counts['CBullDivg_x2']['classic'] += 1
        
        if 'HBearDivg' in analysis_results:
            if "HBearD_gen" in df.columns and pd.notna(df["HBearD_gen"].iloc[i]) and df["HBearD_gen"].iloc[i] == 1:
                markers.append({'Type': 'HBearDivg_Classic', 'Date': current_date, 'Candle_Percent': candle_percent, 'MACD_Percent': macd_percent})
                counts['HBearDivg']['classic'] += 1
        
        if 'HBullDivg' in analysis_results:
            if "HBullD_gen" in df.columns and pd.notna(df["HBullD_gen"].iloc[i]) and df["HBullD_gen"].iloc[i] == 1:
                markers.append({'Type': 'HBullDivg_Classic', 'Date': current_date, 'Candle_Percent': candle_percent, 'MACD_Percent': macd_percent})
                counts['HBullDivg']['classic'] += 1
            if "HBullD_neg_MACD" in df.columns and pd.notna(df["HBullD_neg_MACD"].iloc[i]) and df["HBullD_neg_MACD"].iloc[i] == 1:
                markers.append({'Type': 'HBullDivg_Hidden', 'Date': current_date, 'Candle_Percent': candle_percent, 'MACD_Percent': macd_percent})
                counts['HBullDivg']['hidden'] += 1

    for analysis in counts:
        counts[analysis]['total'] = counts[analysis]['classic'] + counts[analysis]['hidden']

    print("\n=== ANALYSIS SUMMARY ===")
    for analysis_name, count_data in counts.items():
        print(f"{analysis_name}: Classic={count_data['classic']}, Hidden={count_data['hidden']}, Total={count_data['total']}")
    print(f"\nGrand Total markers found: {len(markers)}")
    
    if markers:
        os.makedirs('results', exist_ok=True)
        results_filename = os.path.join('results', filename)
        
        df_markers = pd.DataFrame(markers)
        df_markers.to_csv(results_filename, index=False)
        print(f"\nMarkers exported to {results_filename}")
        
        if openpyxl:
            xlsx_filename = results_filename.replace('.csv', '.xlsx')
            df_xlsx = df_markers.copy()
            df_xlsx['Date'] = pd.to_datetime(df_xlsx['Date']).dt.tz_localize(None)
            
            with pd.ExcelWriter(xlsx_filename, engine='openpyxl') as writer:
                df_xlsx_sorted = df_xlsx.sort_values(['Type', 'Date'])
                df_xlsx_sorted.to_excel(writer, sheet_name='All_Markers', index=False)
                for marker_type in df_xlsx['Type'].unique():
                    subset = df_xlsx[df_xlsx['Type'] == marker_type].sort_values('Date')
                    sheet_name = marker_type.replace('_', ' ')[:31] # Excel sheet name limit
                    subset.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Add Summary sheet
                summary_df = pd.DataFrame.from_dict(counts, orient='index')[['classic', 'hidden', 'total']]
                summary_df.index.name = 'Analysis'
                summary_df.reset_index(inplace=True)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Add grand total row using openpyxl
                worksheet = writer.sheets['Summary']
                grand_total_row = len(summary_df) + 2  # 1-based, after header and data
                worksheet.cell(row=grand_total_row, column=1).value = 'Grand Total'
                worksheet.cell(row=grand_total_row, column=2).value = f'=SUM(B2:B{grand_total_row-1})'
                worksheet.cell(row=grand_total_row, column=3).value = f'=SUM(C2:C{grand_total_row-1})'
                worksheet.cell(row=grand_total_row, column=4).value = f'=SUM(D2:D{grand_total_row-1})'
            
            print(f"Enhanced XLSX exported to {xlsx_filename}")
    
    return counts

def add_markers_to_plotly(fig, df, analysis_results, variant_name):
    marker_positions = set()
    
    marker_styles = {
        'CBullDivg': {'symbol': 'triangle-up', 'color': 'green'},  # ▲ for CBullDivg
        'CBullDivg_x2': {'symbol': 'triangle-down', 'color': 'green'},  # ▼ for CBullDivg_x2
        'HBullDivg': {'symbol': 'diamond', 'color': 'green'},  # ◇ for HBullDivg
        'HBearDivg': {'symbol': 'square', 'color': 'red'}  # □ for HBearDivg
    }

    def add_trace(fig, dates, y_price, y_rsi, y_macd, name, style, direction, div_type, size=12, opacity=1.0):
        legend_group = f"{variant_name}/{div_type}/{direction}"
        legend_group_title = f"{variant_name} - {div_type} - {direction}"
        # Only show legend for the first trace to avoid duplicates
        show_legend = True if name.endswith('Classic') or name.endswith('x2') else False
        fig.add_trace(go.Scatter(x=dates, y=y_price, mode='markers', name=name,
                                 marker=dict(symbol=style['symbol'], color=style['color'], size=size, opacity=opacity, line=dict(width=2)),
                                 legendgroup=legend_group, legendgrouptitle_text=legend_group_title, showlegend=show_legend), row=1, col=1)
        fig.add_trace(go.Scatter(x=dates, y=y_rsi, mode='markers', name=name,
                                 marker=dict(symbol=style['symbol'], color=style['color'], size=size, opacity=opacity, line=dict(width=2)),
                                 legendgroup=legend_group, showlegend=False), row=2, col=1)
        fig.add_trace(go.Scatter(x=dates, y=y_macd, mode='markers', name=name,
                                 marker=dict(symbol=style['symbol'], color=style['color'], size=size, opacity=opacity, line=dict(width=2)),
                                 legendgroup=legend_group, showlegend=False), row=3, col=1)

    df_indexed = df.set_index('date')

    for i in range(len(df)):
        current_date = df['date'].iloc[i]
        
        def get_vals(date_val):
            try:
                row = df_indexed.loc[date_val]
                return row['low'], row['high'], row['RSI'], row['macd_histogram']
            except KeyError:
                return np.nan, np.nan, np.nan, np.nan

        # CBullDivg Classic
        if 'CBullDivg' in analysis_results and "CBullD_gen" in df.columns and df["CBullD_gen"].iloc[i] == 1:
            d1, d2 = pd.to_datetime(df["CBullD_Lower_Low_date_gen"][i]), pd.to_datetime(df["CBullD_Higher_Low_date_gen"][i])
            low1, _, rsi1, macd1 = get_vals(d1)
            low2, _, rsi2, macd2 = get_vals(d2)
            if pd.notna(low1) and pd.notna(low2):
                add_trace(fig, [d1, d2], [low1*0.99, low2*0.99], [rsi1, rsi2], [macd1, macd2],
                          'CBullDivg Classic', marker_styles['CBullDivg'], 'Bullish', 'CBullDivg')
                marker_positions.update([(d1, low1*0.99), (d2, low2*0.99)])
        
        # CBullDivg Hidden
        if 'CBullDivg' in analysis_results and "CBullD_neg_MACD" in df.columns and df["CBullD_neg_MACD"].iloc[i] == 1:
            d1, d2 = pd.to_datetime(df["CBullD_Lower_Low_date_neg_MACD"][i]), pd.to_datetime(df["CBullD_Higher_Low_date_neg_MACD"][i])
            low1, _, rsi1, macd1 = get_vals(d1)
            low2, _, rsi2, macd2 = get_vals(d2)
            if pd.notna(low1) and pd.notna(low2):
                add_trace(fig, [d1, d2], [low1*0.98, low2*0.98], [rsi1, rsi2], [macd1, macd2],
                          'CBullDivg Hidden', marker_styles['CBullDivg'], 'Bullish', 'CBullDivg', size=10, opacity=0.7)
                marker_positions.update([(d1, low1*0.98), (d2, low2*0.98)])

        # CBullDivg_x2
        if 'CBullDivg_x2' in analysis_results and "CBullD_x2_gen" in df.columns and df["CBullD_x2_gen"].iloc[i] == 1:
            d1, d2 = pd.to_datetime(df["CBullD_x2_Lower_Low_date_gen"][i]), pd.to_datetime(df["CBullD_x2_Higher_Low_date_gen"][i])
            low1, _, rsi1, macd1 = get_vals(d1)
            low2, _, rsi2, macd2 = get_vals(d2)
            if pd.notna(low1) and pd.notna(low2):
                add_trace(fig, [d1, d2], [low1*0.97, low2*0.97], [rsi1, rsi2], [macd1, macd2],
                          'CBullDivg x2', marker_styles['CBullDivg_x2'], 'Bullish', 'CBullDivg_x2')
                marker_positions.update([(d1, low1*0.97), (d2, low2*0.97)])
            
        # HBearDivg
        if 'HBearDivg' in analysis_results and "HBearD_gen" in df.columns and df["HBearD_gen"].iloc[i] == 1:
            d1, d2 = pd.to_datetime(df["HBearD_Higher_High_date_gen"][i]), pd.to_datetime(df["HBearD_Lower_High_date_gen"][i])
            _, high1, rsi1, macd1 = get_vals(d1)
            _, high2, rsi2, macd2 = get_vals(d2)
            if pd.notna(high1) and pd.notna(high2):
                add_trace(fig, [d1, d2], [high1*1.01, high2*1.01], [rsi1, rsi2], [macd1, macd2],
                          'HBearDivg Classic', marker_styles['HBearDivg'], 'Bearish', 'HBearDivg')
                marker_positions.update([(d1, high1*1.01), (d2, high2*1.01)])
            
        # HBullDivg Classic
        if 'HBullDivg' in analysis_results and "HBullD_gen" in df.columns and df["HBullD_gen"].iloc[i] == 1:
            d1, d2 = pd.to_datetime(df["HBullD_Lower_Low_date_gen"][i]), pd.to_datetime(df["HBullD_Higher_Low_date_gen"][i])
            low1, _, rsi1, macd1 = get_vals(d1)
            low2, _, rsi2, macd2 = get_vals(d2)
            if pd.notna(low1) and pd.notna(low2):
                add_trace(fig, [d1, d2], [low1*0.96, low2*0.96], [rsi1, rsi2], [macd1, macd2],
                          'HBullDivg Classic', marker_styles['HBullDivg'], 'Bullish', 'HBullDivg')
                marker_positions.update([(d1, low1*0.96), (d2, low2*0.96)])
            
        # HBullDivg Hidden
        if 'HBullDivg' in analysis_results and "HBullD_neg_MACD" in df.columns and df["HBullD_neg_MACD"].iloc[i] == 1:
            d1, d2 = pd.to_datetime(df["HBullD_Lower_Low_date_neg_MACD"][i]), pd.to_datetime(df["HBullD_Higher_Low_date_neg_MACD"][i])
            low1, _, rsi1, macd1 = get_vals(d1)
            low2, _, rsi2, macd2 = get_vals(d2)
            if pd.notna(low1) and pd.notna(low2):
                add_trace(fig, [d1, d2], [low1*0.95, low2*0.95], [rsi1, rsi2], [macd1, macd2],
                          'HBullDivg Hidden', marker_styles['HBullDivg'], 'Bullish', 'HBullDivg', size=10, opacity=0.7)
                marker_positions.update([(d1, low1*0.95), (d2, low2*0.95)])
            
    return marker_positions

def plot_with_plotly(df, analysis_results, counts, df_var2=None, variant2_results=None):
    fig = make_subplots(rows=3, cols=1, shared_xaxes=True,
                        vertical_spacing=0.05,
                        subplot_titles=('Price Chart', 'RSI', 'MACD Histogram'),
                        row_heights=[0.6, 0.2, 0.2])

    fig.add_trace(go.Candlestick(x=df.index, open=df['open'], high=df['high'],
                                 low=df['low'], close=df['close'], name='Price',
                                 increasing_line_color='#44ff44', decreasing_line_color='#ff4444'),
                  row=1, col=1)
    emas = {'EMA_20': 'yellow', 'EMA_50': 'cyan', 'EMA_100': 'magenta', 'EMA_200': 'orange'}
    for ema, color in emas.items():
        fig.add_trace(go.Scatter(x=df.index, y=df[ema], mode='lines', name=ema,
                                 line=dict(color=color, width=1), opacity=0.7), row=1, col=1)
    
    fig.add_trace(go.Scatter(x=df.index, y=df['RSI'], mode='lines', name='RSI',
                             line=dict(color='orange', width=2)), row=2, col=1)

    colors = ['#00FF00' if val >= 0 else '#FF0000' for val in df['macd_histogram']]  # Clear green/red for MACD
    fig.add_trace(go.Bar(x=df.index, y=df['macd_histogram'], name='MACD Hist.',
                         marker_color=colors), row=3, col=1)
    
    df_for_markers = df.reset_index()
    main_positions = add_markers_to_plotly(fig, df_for_markers, analysis_results, 'V1 (Main)')
    
    if df_var2 is not None and variant2_results is not None:
        df_var2_for_markers = df_var2.reset_index()
        variant2_positions = add_markers_to_plotly(fig, df_var2_for_markers, variant2_results, 'V2 (Variant)')
        
        additional = variant2_positions - main_positions
        missing = main_positions - variant2_positions
        
        if additional:
            dates, prices = zip(*additional)
            fig.add_trace(go.Scatter(x=list(dates), y=list(prices), mode='markers', name='V2 Additional',
                                     marker=dict(symbol='circle-open', color='yellow', size=18, line=dict(width=1.5)),
                                     legendgroup='Comparison', legendgrouptitle_text='Comparison'), row=1, col=1)
            fig.add_trace(go.Scatter(x=list(dates), y=[df_for_markers.set_index('date').loc[d]['RSI'] for d in dates], 
                                     mode='markers', name='V2 Additional',
                                     marker=dict(symbol='circle-open', color='yellow', size=18, line=dict(width=1.5)),
                                     legendgroup='Comparison', showlegend=False), row=2, col=1)
            fig.add_trace(go.Scatter(x=list(dates), y=[df_for_markers.set_index('date').loc[d]['macd_histogram'] for d in dates], 
                                     mode='markers', name='V2 Additional',
                                     marker=dict(symbol='circle-open', color='yellow', size=18, line=dict(width=1.5)),
                                     legendgroup='Comparison', showlegend=False), row=3, col=1)
        if missing:
            dates, prices = zip(*missing)
            fig.add_trace(go.Scatter(x=list(dates), y=list(prices), mode='markers', name='V2 Missing',
                                     marker=dict(symbol='circle-open', color='blue', size=18, line=dict(width=1.5)),
                                     legendgroup='Comparison', legendgrouptitle_text='Comparison'), row=1, col=1)
            fig.add_trace(go.Scatter(x=list(dates), y=[df_for_markers.set_index('date').loc[d]['RSI'] for d in dates], 
                                     mode='markers', name='V2 Missing',
                                     marker=dict(symbol='circle-open', color='blue', size=18, line=dict(width=1.5)),
                                     legendgroup='Comparison', showlegend=False), row=2, col=1)
            fig.add_trace(go.Scatter(x=list(dates), y=[df_for_markers.set_index('date').loc[d]['macd_histogram'] for d in dates], 
                                     mode='markers', name='V2 Missing',
                                     marker=dict(symbol='circle-open', color='blue', size=18, line=dict(width=1.5)),
                                     legendgroup='Comparison', showlegend=False), row=3, col=1)

    fig.update_layout(
        title_text='Technical Analysis with Divergence Markers',
        xaxis_rangeslider_visible=False,
        template='plotly_dark',
        legend=dict(traceorder='grouped', groupclick='toggleitem'),
        height=900
    )
    fig.update_yaxes(title_text="Price", row=1, col=1)
    fig.update_yaxes(title_text="RSI", range=[0, 100], row=2, col=1)
    fig.update_yaxes(title_text="MACD", row=3, col=1)

    count_lines = [f"<b>{name}</b>: C={data['classic']}, H={data['hidden']}, T={data['total']}" for name, data in counts.items()]
    grand_total = sum(data['total'] for data in counts.values())
    count_lines.append(f"<b>Grand Total: {grand_total}</b>")
    count_text = "<br>".join(count_lines)
    
    fig.add_annotation(
        text=count_text, align='left', showarrow=False, xref='paper', yref='paper',
        x=0.01, y=0.99, bgcolor="rgba(0,0,0,0.7)", bordercolor="white", borderwidth=1
    )

    os.makedirs('results', exist_ok=True)
    plot_filename = os.path.join('results', 'plot.html')
    fig.write_html(plot_filename)
    print(f"\nInteractive plot saved to: {plot_filename}")
    fig.show()

def run_doe_analysis(df, doe_params_file="doe_parameters_example.csv"):
    print(f"Loading DOE parameters from {doe_params_file}...")
    try:
        doe_params = pd.read_csv(doe_params_file)
        if not all(col in doe_params.columns for col in ['candle_percent', 'macd_percent']):
            raise ValueError("DOE parameters file must contain 'candle_percent' and 'macd_percent' columns")
    except Exception as e:
        print(f"Error loading DOE parameters: {e}. Exiting DOE analysis.")
        return

    window = 5
    all_counts = []
    all_markers = []

    for idx, row in doe_params.iterrows():
        candle_tol = row['candle_percent']
        macd_tol = row['macd_percent']
        print(f"\nRunning DOE iteration {idx+1}: candle_tol={candle_tol}, macd_tol={macd_tol}")
        
        df_copy = df.copy()
        analysis_results = run_analysis(df_copy, 'e', window, candle_tol, macd_tol)
        
        output_filename = f"doe_markers_candle{candle_tol}_macd{macd_tol}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv"
        counts = export_markers_to_csv(df_copy, output_filename, analysis_results, candle_tol, macd_tol)
        
        for analysis, count_data in counts.items():
            all_counts.append({
                'Candle_Percent': candle_tol,
                'MACD_Percent': macd_tol,
                'Analysis': analysis,
                'Classic': count_data['classic'],
                'Hidden': count_data['hidden'],
                'Total': count_data['total']
            })
        
        df_markers = pd.read_csv(os.path.join('results', output_filename))
        df_markers['Candle_Percent'] = candle_tol
        df_markers['MACD_Percent'] = macd_tol
        all_markers.append(df_markers)

    if all_counts:
        doe_summary_df = pd.DataFrame(all_counts)
        os.makedirs('results', exist_ok=True)
        summary_filename = os.path.join('results', f"doe_summary_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        with pd.ExcelWriter(summary_filename, engine='openpyxl') as writer:
            doe_summary_df.to_excel(writer, sheet_name='DOE_Summary', index=False)
            pd.concat(all_markers).to_excel(writer, sheet_name='All_Markers', index=False)
            
            worksheet = writer.sheets['DOE_Summary']
            grand_total_row = len(doe_summary_df) + 2
            worksheet.cell(row=grand_total_row, column=1).value = 'Grand Total'
            worksheet.cell(row=grand_total_row, column=4).value = f'=SUM(D2:D{grand_total_row-1})'
            worksheet.cell(row=grand_total_row, column=5).value = f'=SUM(E2:E{grand_total_row-1})'
            worksheet.cell(row=grand_total_row, column=6).value = f'=SUM(F2:F{grand_total_row-1})'
        
        print(f"DOE summary exported to {summary_filename}")

if __name__ == "__main__":
    csv_file_path = get_input_file()
    if not csv_file_path:
        print("No file selected. Exiting.")
        exit()

    analysis_type = get_analysis_type()

    if analysis_type != 'f':
        candle_percent, macd_percent, variant2 = get_analysis_parameters()

    print(f"Loading data from: {csv_file_path}")
    if csv_file_path.lower().endswith('.parquet'):
        df = pd.read_parquet(csv_file_path)
    else:
        df = pd.read_csv(csv_file_path, low_memory=False)
    
    print(f"Loaded {len(df)} rows")
    df['date'] = pd.to_datetime(df['date'])
    df.sort_values(by='date', inplace=True)
    df.reset_index(drop=True, inplace=True)

    print("Initializing indicators...")
    Initialize_RSI_EMA_MACD(df)
    Local_Max_Min(df)

    window = 5

    if analysis_type != 'f':
        print(f"Running analysis with parameters: window={window}, candle_tol={candle_percent}, macd_tol={macd_percent}")
        analysis_results = run_analysis(df, analysis_type, window, candle_percent, macd_percent)
        
        output_filename = f"markers_output_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv"
        counts = export_markers_to_csv(df, output_filename, analysis_results, candle_percent, macd_percent)
        
        df_var2, variant2_results = None, None
        if variant2:
            print(f"\nRunning second variant: candle={variant2[0]}, macd={variant2[1]}")
            df_var2 = df.copy() 
            variant2_results = run_analysis(df_var2, analysis_type, window, variant2[0], variant2[1])
            export_markers_to_csv(df_var2, f"variant2_{variant2[0]}_{variant2[1]}.csv", variant2_results, variant2[0], variant2[1])
        
        df.set_index('date', inplace=True)
        if df_var2 is not None:
            df_var2.set_index('date', inplace=True)
            
        plot_with_plotly(df, analysis_results, counts, df_var2, variant2_results)
    else:
        run_doe_analysis(df)