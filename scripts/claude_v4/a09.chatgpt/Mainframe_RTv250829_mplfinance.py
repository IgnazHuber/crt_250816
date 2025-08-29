# Mainframe_RTv250829_mplfinance.py
# Vollständige korrigierte Version mit präziser Legendengruppierung
# - 1 Gruppe pro Variante (V1/V2)
# - 1 Eintrag pro (Divergenztyp, Richtung) innerhalb der Variante
# - Classic & Hidden unter demselben Legendeneintrag (gemeinsame legendgroup)
# - Comparison (V2 Additional/Missing) separat
# - Keine Features entfernt; nur Fixes/Ergänzungen

import os
import glob
import subprocess
from datetime import datetime

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# optionale Excel-Unterstützung
try:
    import openpyxl  # noqa: F401
    _HAS_XLSX = True
except Exception:
    _HAS_XLSX = False

# --- externe (bestehende) Module: NICHT ändern ---
from CBullDivg_Analysis_vectorized import CBullDivg_analysis
from CBullDivg_x2_analysis_vectorized import CBullDivg_x2_analysis
from HBearDivg_analysis_vectorized import HBearDivg_analysis
from HBullDivg_analysis_vectorized import HBullDivg_analysis
from Initialize_RSI_EMA_MACD import Initialize_RSI_EMA_MACD
from Local_Maximas_Minimas import Local_Max_Min


# -------------------------------------------------
# Datei-Auswahl (PowerShell Dialog mit Fallback)
# -------------------------------------------------
def get_input_file():
    """PowerShell-OpenFileDialog; Fallback: Konsole."""
    try:
        ps_command = r'''
Add-Type -AssemblyName System.Windows.Forms
$dlg = New-Object System.Windows.Forms.OpenFileDialog
$dlg.Filter = "CSV/Parquet (*.csv;*.parquet)|*.csv;*.parquet|CSV (*.csv)|*.csv|Parquet (*.parquet)|*.parquet|All files (*.*)|*.*"
$dlg.Title = "Select CSV/Parquet input file"
$dlg.InitialDirectory = "C:\Projekte\crt_250816\data\raw"
if ($dlg.ShowDialog() -eq [System.Windows.Forms.DialogResult]::OK) { Write-Output $dlg.FileName }
'''
        result = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_command],
            capture_output=True, text=True, timeout=60
        )
        path = (result.stdout or "").strip()
        if result.returncode == 0 and path:
            return path
    except Exception:
        pass
    return get_input_file_console()


def get_input_file_console():
    default_file = r"C:\Projekte\crt_250816\data\raw\btc_1day_candlesticks_all.csv"
    print(f"\nDefault file: {default_file}")
    print("Press Enter to use default, or choose from available files:")

    candidates = []
    search_patterns = [
        "*.csv", "*.parquet",
        "../*/*.csv", "../*/*.parquet",
        "../../data/raw/*.csv", "../../data/raw/*.parquet",
        "../../data/*.csv", "../../data/*.parquet"
    ]
    for pat in search_patterns:
        candidates.extend(glob.glob(pat))

    for i, f in enumerate(candidates, 1):
        print(f"{i}: {f}")
    print(f"{len(candidates) + 1}: Enter custom path")
    print(f"0 or Enter: Use default ({default_file})")

    while True:
        try:
            choice = input(f"\nSelect file (0-{len(candidates) + 1}): ").strip()
            if choice in ("", "0"):
                return default_file
            if choice.isdigit():
                idx = int(choice)
                if 1 <= idx <= len(candidates):
                    return candidates[idx - 1]
                if idx == len(candidates) + 1:
                    custom = input("Enter full path (or Enter for default): ").strip()
                    return custom if custom else default_file
            print("Invalid choice. Try again.")
        except (EOFError, KeyboardInterrupt):
            return default_file


# -------------------------------------------------
# Parameter/Typ-Auswahl
# -------------------------------------------------
def get_analysis_parameters():
    import sys
    # CLI: script <type> <candle> <macd>
    if len(sys.argv) > 2:
        try:
            candle_percent = float(sys.argv[2])
            macd_percent = float(sys.argv[3]) if len(sys.argv) > 3 else 3.25
            return candle_percent, macd_percent, None
        except Exception:
            pass

    try:
        print("\nEnter analysis parameters (or press Enter for defaults):")
        s = input("Enter Candle tolerance % (default: 0.1): ").strip()
        candle_percent = 0.1 if s == "" else float(s)
        s = input("Enter MACD tolerance % (default: 3.25): ").strip()
        macd_percent = 3.25 if s == "" else float(s)

        print("\nOptional second variant for comparison:")
        s1 = input("Enter second Candle tolerance % (leave empty for none): ").strip()
        s2 = input("Enter second MACD tolerance % (leave empty for none): ").strip()

        variant2 = None
        if s1 or s2:
            try:
                c2 = float(s1) if s1 else candle_percent
                m2 = float(s2) if s2 else macd_percent
                variant2 = (c2, m2)
            except Exception as e:
                print(f"Invalid second variant: {e}")
                variant2 = None

        return candle_percent, macd_percent, variant2
    except (EOFError, KeyboardInterrupt):
        return 0.1, 3.25, None


def get_analysis_type():
    import sys
    if len(sys.argv) > 1:
        c = (sys.argv[1] or "").lower().strip()
        if c in list("abcdef"):
            return c

    try:
        print("\nSelect analysis type:")
        print("a: CBullDivg_Analysis (Classic Bullish Divergence)")
        print("b: BullDivg_x2_analysis (Extended Bullish Divergence)")
        print("c: HBearDivg_analysis (Hidden Bearish Divergence)")
        print("d: HBullDivg_analysis (Hidden Bullish Divergence)")
        print("e: All analyses (a-d)")
        print("f: DOE (Design of Experiments)")
        while True:
            c = input("\nEnter your choice (a-f): ").lower().strip()
            if c in list("abcdef"):
                return c
            print("Invalid choice. Please enter a, b, c, d, e, or f.")
    except (EOFError, KeyboardInterrupt):
        return "e"


# -------------------------------------------------
# Analysen laufen lassen (ruft vorhandene Module)
# -------------------------------------------------
def run_analysis(df, analysis_type, window=5, candle_tol=0.1, macd_tol=3.25):
    results = {}
    if analysis_type in ['a', 'e', 'f']:
        CBullDivg_analysis(df, window, candle_tol, macd_tol)
        results['CBullDivg'] = True
    if analysis_type in ['b', 'e', 'f']:
        CBullDivg_x2_analysis(df, window, candle_tol, macd_tol)
        results['CBullDivg_x2'] = True
    if analysis_type in ['c', 'e', 'f']:
        HBearDivg_analysis(df, window, candle_tol, macd_tol)
        results['HBearDivg'] = True
    if analysis_type in ['d', 'e', 'f']:
        HBullDivg_analysis(df, window, candle_tol, macd_tol)
        results['HBullDivg'] = True
    return results


# -------------------------------------------------
# Marker-Export (CSV/XLSX mit Summary)
# -------------------------------------------------
def export_markers_to_csv(df, filename, analysis_results, candle_percent, macd_percent):
    os.makedirs('results', exist_ok=True)
    markers = []
    counts = {k: {'classic': 0, 'hidden': 0, 'total': 0}
              for k, v in analysis_results.items() if v}

    def _append(t, dt):
        markers.append({
            'Type': t,
            'Date': dt,
            'Candle_Percent': candle_percent,
            'MACD_Percent': macd_percent
        })

    for i in range(len(df)):
        dt = df['date'].iloc[i]
        # Classic Bullish
        if analysis_results.get('CBullDivg', False):
            if 'CBullD_gen' in df.columns and pd.notna(df.at[i, 'CBullD_gen']) and df.at[i, 'CBullD_gen'] == 1:
                _append('CBullDivg_Classic', dt)
                counts['CBullDivg']['classic'] += 1
            if 'CBullD_neg_MACD' in df.columns and pd.notna(df.at[i, 'CBullD_neg_MACD']) and df.at[i, 'CBullD_neg_MACD'] == 1:
                _append('CBullDivg_Hidden', dt)
                counts['CBullDivg']['hidden'] += 1

        # x2
        if analysis_results.get('CBullDivg_x2', False):
            if 'CBullD_x2_gen' in df.columns and pd.notna(df.at[i, 'CBullD_x2_gen']) and df.at[i, 'CBullD_x2_gen'] == 1:
                _append('CBullDivg_x2_Classic', dt)
                counts['CBullDivg_x2']['classic'] += 1

        # Hidden Bearish
        if analysis_results.get('HBearDivg', False):
            if 'HBearD_gen' in df.columns and pd.notna(df.at[i, 'HBearD_gen']) and df.at[i, 'HBearD_gen'] == 1:
                _append('HBearDivg_Classic', dt)
                counts['HBearDivg']['classic'] += 1

        # Hidden Bullish
        if analysis_results.get('HBullDivg', False):
            if 'HBullD_gen' in df.columns and pd.notna(df.at[i, 'HBullD_gen']) and df.at[i, 'HBullD_gen'] == 1:
                _append('HBullDivg_Classic', dt)
                counts['HBullDivg']['classic'] += 1
            if 'HBullD_neg_MACD' in df.columns and pd.notna(df.at[i, 'HBullD_neg_MACD']) and df.at[i, 'HBullD_neg_MACD'] == 1:
                _append('HBullDivg_Hidden', dt)
                counts['HBullDivg']['hidden'] += 1

    # Totale
    for k in counts:
        counts[k]['total'] = counts[k]['classic'] + counts[k]['hidden']

    print("\n=== ANALYSIS SUMMARY ===")
    for name, c in counts.items():
        print(f"{name}: Classic={c['classic']}, Hidden={c['hidden']}, Total={c['total']}")
    print(f"\nGrand Total markers found: {len(markers)}")

    # CSV
    csv_path = os.path.join('results', filename)
    pd.DataFrame(markers).to_csv(csv_path, index=False)
    print(f"\nMarkers exported to {csv_path}")

    # XLSX (optional)
    if _HAS_XLSX:
        xlsx_path = csv_path.replace('.csv', '.xlsx')
        df_x = pd.DataFrame(markers).copy()
        if not df_x.empty:
            df_x['Date'] = pd.to_datetime(df_x['Date']).dt.tz_localize(None)
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            # alle Marker (sortiert)
            df_x.sort_values(['Type', 'Date']).to_excel(writer, sheet_name='All_Markers', index=False)
            # pro Typ eigenes Blatt
            for t in df_x['Type'].dropna().unique():
                df_x[df_x['Type'] == t].sort_values('Date').to_excel(
                    writer, sheet_name=str(t).replace('_', ' ')[:31], index=False
                )
            # Summary
            s = (pd.DataFrame.from_dict(counts, orient='index')[['classic', 'hidden', 'total']]
                 .rename_axis('Analysis').reset_index())
            s.to_excel(writer, sheet_name='Summary', index=False)
            ws = writer.sheets['Summary']
            grand_row = len(s) + 2
            ws.cell(row=grand_row, column=1).value = 'Grand Total'
            ws.cell(row=grand_row, column=2).value = f'=SUM(B2:B{grand_row - 1})'
            ws.cell(row=grand_row, column=3).value = f'=SUM(C2:C{grand_row - 1})'
            ws.cell(row=grand_row, column=4).value = f'=SUM(D2:D{grand_row - 1})'
        print(f"Enhanced XLSX exported to {xlsx_path}")

    return counts


# -------------------------------------------------
# Plot-Hilfen
# -------------------------------------------------
def _safe_get_vals(df_idx, dt):
    """Liefert Low, High, RSI, MACD für gegebenes Datum; NaN falls nicht vorhanden."""
    try:
        row = df_idx.loc[dt]
        return row.get('low', np.nan), row.get('high', np.nan), row.get('RSI', np.nan), row.get('macd_histogram', np.nan)
    except Exception:
        return np.nan, np.nan, np.nan, np.nan


def add_markers_to_plotly(fig, df, analysis_results, variant_name):
    """
    Zeichnet Marker in 3 Subplots (Price/RSI/MACD).
    Legendengruppierung exakt wie gefordert:
      - legendgroup = f"{variant}|{div_type}|{direction}"
      - Pro Gruppe nur EIN sichtbarer Legenden-Eintrag (Proxy-Trace mit visible='legendonly')
      - Classic/Hidden-Traces: showlegend=False, gleiche legendgroup
    Rückgabe: Set aus (date, y_price) zur Varianten-Differenz.
    """
    positions = set()

    marker_styles = {
        'CBullDivg':    {'symbol': 'triangle-up',   'color': 'green'},
        'CBullDivg_x2': {'symbol': 'triangle-down', 'color': 'green'},
        'HBullDivg':    {'symbol': 'diamond',       'color': 'green'},
        'HBearDivg':    {'symbol': 'square',        'color': 'red'}
    }

    # Hilfsspeicher: welche Gruppen haben bereits echte Daten bekommen?
    group_seen = set()

    def _group_id(div_type, direction):
        return f"{variant_name}|{div_type}|{direction}"

    def _proxy_name(div_type, direction):
        # Gewünscht: "V1 (Main) - CBullDivg - Bullish"
        return f"{variant_name} - {div_type} - {direction}"

    def _add_proxy_legend_entry(div_type, direction):
        gid = _group_id(div_type, direction)
        if gid in group_seen:
            # Proxy bereits vorhanden
            return
        # Dummy-Trace nur für die Legende (legendonly), gruppiert & mit Titel = Variante
        fig.add_trace(
            go.Scatter(
                x=[df['date'].iloc[0] if len(df) else None],
                y=[np.nan],  # nicht gezeichnet
                mode='markers',
                name=_proxy_name(div_type, direction),
                marker=dict(
                    symbol=marker_styles[div_type]['symbol'],
                    size=12,
                    line=dict(width=2),
                ),
                visible='legendonly',
                legendgroup=_group_id(div_type, direction),
                legendgrouptitle_text=variant_name,
                showlegend=True,
                hoverinfo='skip'
            ),
            row=1, col=1
        )
        group_seen.add(gid)

    def _add_pair(dates, y_price, y_rsi, y_macd, div_type, direction, size, opacity):
        gid = _group_id(div_type, direction)
        common_marker = dict(symbol=marker_styles[div_type]['symbol'],
                             color=marker_styles[div_type]['color'],
                             size=size, opacity=opacity, line=dict(width=2))
        # Preis
        fig.add_trace(
            go.Scatter(x=dates, y=y_price, mode='markers',
                       name=f"{div_type} - {direction}",
                       marker=common_marker,
                       legendgroup=gid,
                       showlegend=False),
            row=1, col=1
        )
        # RSI
        fig.add_trace(
            go.Scatter(x=dates, y=y_rsi, mode='markers',
                       name=f"{div_type} - {direction}",
                       marker=common_marker,
                       legendgroup=gid,
                       showlegend=False),
            row=2, col=1
        )
        # MACD
        fig.add_trace(
            go.Scatter(x=dates, y=y_macd, mode='markers',
                       name=f"{div_type} - {direction}",
                       marker=common_marker,
                       legendgroup=gid,
                       showlegend=False),
            row=3, col=1
        )
        # Sicherstellen, dass genau EIN Legendeneintrag für diese Gruppe existiert
        _add_proxy_legend_entry(div_type, direction)

    df_idx = df.set_index('date')

    for i in range(len(df)):
        # Classic Bullish
        if analysis_results.get('CBullDivg', False) and 'CBullD_gen' in df.columns and df.at[i, 'CBullD_gen'] == 1:
            d1 = pd.to_datetime(df.at[i, 'CBullD_Lower_Low_date_gen'])
            d2 = pd.to_datetime(df.at[i, 'CBullD_Higher_Low_date_gen'])
            low1, _, rsi1, macd1 = _safe_get_vals(df_idx, d1)
            low2, _, rsi2, macd2 = _safe_get_vals(df_idx, d2)
            if pd.notna(low1) and pd.notna(low2):
                _add_pair([d1, d2], [low1 * 0.99, low2 * 0.99], [rsi1, rsi2], [macd1, macd2],
                          'CBullDivg', 'Bullish', size=12, opacity=1.0)
                positions.update([(d1, low1 * 0.99), (d2, low2 * 0.99)])

        # Hidden Bullish (neg MACD)
        if analysis_results.get('CBullDivg', False) and 'CBullD_neg_MACD' in df.columns and df.at[i, 'CBullD_neg_MACD'] == 1:
            d1 = pd.to_datetime(df.at[i, 'CBullD_Lower_Low_date_neg_MACD'])
            d2 = pd.to_datetime(df.at[i, 'CBullD_Higher_Low_date_neg_MACD'])
            low1, _, rsi1, macd1 = _safe_get_vals(df_idx, d1)
            low2, _, rsi2, macd2 = _safe_get_vals(df_idx, d2)
            if pd.notna(low1) and pd.notna(low2):
                _add_pair([d1, d2], [low1 * 0.98, low2 * 0.98], [rsi1, rsi2], [macd1, macd2],
                          'CBullDivg', 'Bullish', size=10, opacity=0.7)
                positions.update([(d1, low1 * 0.98), (d2, low2 * 0.98)])

        # x2 Bullish
        if analysis_results.get('CBullDivg_x2', False) and 'CBullD_x2_gen' in df.columns and df.at[i, 'CBullD_x2_gen'] == 1:
            d1 = pd.to_datetime(df.at[i, 'CBullD_x2_Lower_Low_date_gen'])
            d2 = pd.to_datetime(df.at[i, 'CBullD_x2_Higher_Low_date_gen'])
            low1, _, rsi1, macd1 = _safe_get_vals(df_idx, d1)
            low2, _, rsi2, macd2 = _safe_get_vals(df_idx, d2)
            if pd.notna(low1) and pd.notna(low2):
                _add_pair([d1, d2], [low1 * 0.97, low2 * 0.97], [rsi1, rsi2], [macd1, macd2],
                          'CBullDivg_x2', 'Bullish', size=12, opacity=1.0)
                positions.update([(d1, low1 * 0.97), (d2, low2 * 0.97)])

        # Hidden Bearish
        if analysis_results.get('HBearDivg', False) and 'HBearD_gen' in df.columns and df.at[i, 'HBearD_gen'] == 1:
            d1 = pd.to_datetime(df.at[i, 'HBearD_Higher_High_date_gen'])
            d2 = pd.to_datetime(df.at[i, 'HBearD_Lower_High_date_gen'])
            _, high1, rsi1, macd1 = _safe_get_vals(df_idx, d1)
            _, high2, rsi2, macd2 = _safe_get_vals(df_idx, d2)
            if pd.notna(high1) and pd.notna(high2):
                _add_pair([d1, d2], [high1 * 1.01, high2 * 1.01], [rsi1, rsi2], [macd1, macd2],
                          'HBearDivg', 'Bearish', size=12, opacity=1.0)
                positions.update([(d1, high1 * 1.01), (d2, high2 * 1.01)])

        # Hidden Bullish (gen)
        if analysis_results.get('HBullDivg', False) and 'HBullD_gen' in df.columns and df.at[i, 'HBullD_gen'] == 1:
            d1 = pd.to_datetime(df.at[i, 'HBullD_Lower_Low_date_gen'])
            d2 = pd.to_datetime(df.at[i, 'HBullD_Higher_Low_date_gen'])
            low1, _, rsi1, macd1 = _safe_get_vals(df_idx, d1)
            low2, _, rsi2, macd2 = _safe_get_vals(df_idx, d2)
            if pd.notna(low1) and pd.notna(low2):
                _add_pair([d1, d2], [low1 * 0.96, low2 * 0.96], [rsi1, rsi2], [macd1, macd2],
                          'HBullDivg', 'Bullish', size=12, opacity=1.0)
                positions.update([(d1, low1 * 0.96), (d2, low2 * 0.96)])

        # Hidden Bullish (neg MACD)
        if analysis_results.get('HBullDivg', False) and 'HBullD_neg_MACD' in df.columns and df.at[i, 'HBullD_neg_MACD'] == 1:
            d1 = pd.to_datetime(df.at[i, 'HBullD_Lower_Low_date_neg_MACD'])
            d2 = pd.to_datetime(df.at[i, 'HBullD_Higher_Low_date_neg_MACD'])
            low1, _, rsi1, macd1 = _safe_get_vals(df_idx, d1)
            low2, _, rsi2, macd2 = _safe_get_vals(df_idx, d2)
            if pd.notna(low1) and pd.notna(low2):
                _add_pair([d1, d2], [low1 * 0.95, low2 * 0.95], [rsi1, rsi2], [macd1, macd2],
                          'HBullDivg', 'Bullish', size=10, opacity=0.7)
                positions.update([(d1, low1 * 0.95), (d2, low2 * 0.95)])

    return positions


def plot_with_plotly(df_main, analysis_results, counts_main, df_var2=None, variant2_results=None):
    fig = make_subplots(
        rows=3, cols=1, shared_xaxes=True, vertical_spacing=0.05,
        subplot_titles=('Price', 'RSI', 'MACD Histogram'),
        row_heights=[0.4, 0.3, 0.3]
    )

    # Candles
    fig.add_trace(
        go.Candlestick(
            x=df_main.index, open=df_main['open'], high=df_main['high'],
            low=df_main['low'], close=df_main['close'], name='Price',
            increasing_line_color='#44ff44', decreasing_line_color='#ff4444'
        ),
        row=1, col=1
    )

    # EMAs (vorausgesetzt von Initialize_RSI_EMA_MACD)
    ema_defs = {'EMA_20': 'yellow', 'EMA_50': 'cyan', 'EMA_100': 'magenta', 'EMA_200': 'orange'}
    for col, _col_color in ema_defs.items():
        if col in df_main.columns:
            fig.add_trace(
                go.Scatter(x=df_main.index, y=df_main[col], mode='lines', name=col,
                           line=dict(width=1), opacity=0.8),
                row=1, col=1
            )

    # RSI
    if 'RSI' in df_main.columns:
        fig.add_trace(
            go.Scatter(x=df_main.index, y=df_main['RSI'], mode='lines', name='RSI', line=dict(width=2)),
            row=2, col=1
        )

    # MACD-Histogramm
    macd_vals = df_main['macd_histogram'] if 'macd_histogram' in df_main.columns else pd.Series(0, index=df_main.index)
    colors = ['#00FF00' if v >= 0 else '#FF0000' for v in macd_vals]
    fig.add_trace(
        go.Bar(x=df_main.index, y=macd_vals, name='MACD Hist.', marker_color=colors),
        row=3, col=1
    )

    # Marker (Variante 1)
    df_m_for_markers = df_main.reset_index().rename(columns={'index': 'date'})
    main_positions = add_markers_to_plotly(fig, df_m_for_markers, analysis_results, 'V1 (Main)')

    # Marker (Variante 2) + Vergleich Outlines
    if df_var2 is not None and variant2_results is not None:
        df_v2_for_markers = df_var2.reset_index().rename(columns={'index': 'date'})
        v2_positions = add_markers_to_plotly(fig, df_v2_for_markers, variant2_results, 'V2 (Variant)')

        additional = v2_positions - main_positions
        missing = main_positions - v2_positions

        def _add_outline(dates_prices, name, color):
            if not dates_prices:
                return
            dates, prices = zip(*dates_prices)
            rsi_vals = [df_m_for_markers.set_index('date').get('RSI', pd.Series(dtype=float)).get(d, np.nan) for d in dates]
            macd_vals_local = [df_m_for_markers.set_index('date').get('macd_histogram', pd.Series(dtype=float)).get(d, np.nan) for d in dates]

            mk = dict(symbol='circle-open', color=color, size=18, line=dict(width=1.5))
            # Preis
            fig.add_trace(go.Scatter(x=list(dates), y=list(prices), mode='markers',
                                     name=name, marker=mk,
                                     legendgroup=f"Comparison|{name}",
                                     legendgrouptitle_text="Comparison",
                                     showlegend=True),
                          row=1, col=1)
            # RSI
            fig.add_trace(go.Scatter(x=list(dates), y=rsi_vals, mode='markers',
                                     name=name, marker=mk,
                                     legendgroup=f"Comparison|{name}",
                                     showlegend=False),
                          row=2, col=1)
            # MACD
            fig.add_trace(go.Scatter(x=list(dates), y=macd_vals_local, mode='markers',
                                     name=name, marker=mk,
                                     legendgroup=f"Comparison|{name}",
                                     showlegend=False),
                          row=3, col=1)

        _add_outline(additional, 'V2 Additional', 'yellow')
        _add_outline(missing, 'V2 Missing', 'blue')

    # Layout/Legend
    fig.update_layout(
        title_text='Technical Analysis with Divergence Markers',
        xaxis_rangeslider_visible=False,
        template='plotly_dark',
        legend=dict(
            traceorder='grouped',
            groupclick='togglegroup',  # Klick toggelt alle Traces der Gruppe (Classic+Hidden gemeinsam)
            itemclick='toggle'         # einzelner Eintrag (Proxy/Comparison) toggelt normal
        ),
        height=1200
    )
    fig.update_yaxes(title_text="Price", row=1, col=1)
    fig.update_yaxes(title_text="RSI", range=[0, 100], row=2, col=1)
    fig.update_yaxes(title_text="MACD", row=3, col=1)

    # Zählerbox
    lines = [f"<b>{name}</b>: C={d['classic']}, H={d['hidden']}, T={d['total']}" for name, d in counts_main.items()]
    gtotal = sum(d['total'] for d in counts_main.values())
    lines.append(f"<b>Grand Total: {gtotal}</b>")
    fig.add_annotation(text="<br>".join(lines), align='left', showarrow=False,
                       xref='paper', yref='paper', x=0.01, y=0.99,
                       bgcolor="rgba(0,0,0,0.7)", bordercolor="white", borderwidth=1)

    # speichern & anzeigen
    os.makedirs('results', exist_ok=True)
    out_html = os.path.join('results', 'plot.html')
    fig.write_html(out_html)
    print(f"\nInteractive plot saved to: {out_html}")
    fig.show()


# -------------------------------------------------
# DOE
# -------------------------------------------------
def run_doe_analysis(df, doe_params_file="doe_parameters_example.csv"):
    print(f"Loading DOE parameters from {doe_params_file}...")
    try:
        prm = pd.read_csv(doe_params_file)
        required = {'candle_percent', 'macd_percent'}
        if not required.issubset(set(prm.columns)):
            raise ValueError(f"DOE parameters file must contain columns: {sorted(required)}")
    except Exception as e:
        print(f"Error loading DOE parameters: {e}. Exiting DOE analysis.")
        return

    window = 5
    all_counts_rows = []
    all_marker_frames = []

    for idx, row in prm.iterrows():
        candle_tol = float(row['candle_percent'])
        macd_tol = float(row['macd_percent'])
        print(f"\nRunning DOE iteration {idx + 1}: candle_tol={candle_tol}, macd_tol={macd_tol}")

        df_copy = df.copy()
        res = run_analysis(df_copy, 'e', window, candle_tol, macd_tol)

        out_csv = f"doe_markers_candle{candle_tol}_macd{macd_tol}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        counts = export_markers_to_csv(df_copy, out_csv, res, candle_tol, macd_tol)

        for a_name, c_dict in counts.items():
            all_counts_rows.append({
                'Candle_Percent': candle_tol,
                'MACD_Percent': macd_tol,
                'Analysis': a_name,
                'Classic': c_dict['classic'],
                'Hidden': c_dict['hidden'],
                'Total': c_dict['total']
            })

        try:
            df_m = pd.read_csv(os.path.join('results', out_csv))
            df_m['Candle_Percent'] = candle_tol
            df_m['MACD_Percent'] = macd_tol
            all_marker_frames.append(df_m)
        except Exception:
            pass

    if all_counts_rows:
        os.makedirs('results', exist_ok=True)
        xlsx = os.path.join('results', f"doe_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if _HAS_XLSX:
            with pd.ExcelWriter(xlsx, engine='openpyxl') as writer:
                pd.DataFrame(all_counts_rows).to_excel(writer, sheet_name='DOE_Summary', index=False)
                if all_marker_frames:
                    pd.concat(all_marker_frames, ignore_index=True).to_excel(writer, sheet_name='All_Markers', index=False)
                ws = writer.sheets['DOE_Summary']
                n = len(all_counts_rows) + 2
                ws.cell(row=n, column=1).value = 'Grand Total'
                ws.cell(row=n, column=4).value = f'=SUM(D2:D{n - 1})'
                ws.cell(row=n, column=5).value = f'=SUM(E2:E{n - 1})'
                ws.cell(row=n, column=6).value = f'=SUM(F2:F{n - 1})'
            print(f"DOE summary exported to {xlsx}")
        else:
            # CSV-Fallback, wenn kein openpyxl vorhanden ist
            csv = xlsx.replace('.xlsx', '.csv')
            pd.DataFrame(all_counts_rows).to_csv(csv, index=False)
            print(f"DOE summary exported to {csv}")


# -------------------------------------------------
# MAIN
# -------------------------------------------------
if __name__ == "__main__":
    path = get_input_file()
    if not path:
        print("No file selected. Exiting.")
        raise SystemExit(0)

    atype = get_analysis_type()

    # Parameter nur, wenn nicht DOE
    if atype != 'f':
        candle_percent, macd_percent, variant2 = get_analysis_parameters()

    print(f"Loading data from: {path}")
    if path.lower().endswith('.parquet'):
        df = pd.read_parquet(path)
    else:
        df = pd.read_csv(path, low_memory=False)

    print(f"Loaded {len(df)} rows")

    # Spalte 'date' sicherstellen
    if 'date' not in df.columns and 'timestamp' in df.columns:
        df['date'] = df['timestamp']

    df['date'] = pd.to_datetime(df['date'], utc=True, errors='coerce')
    df = df.dropna(subset=['date']).sort_values('date').reset_index(drop=True)

    print("Initializing indicators...")
    Initialize_RSI_EMA_MACD(df)
    Local_Max_Min(df)

    window = 5

    if atype != 'f':
        print(f"Running analysis with parameters: window={window}, candle_tol={candle_percent}, macd_tol={macd_percent}")
        analysis_results = run_analysis(df, atype, window, candle_percent, macd_percent)
        out_csv = f"markers_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        counts = export_markers_to_csv(df, out_csv, analysis_results, candle_percent, macd_percent)

        df2 = None
        res2 = None
        if variant2:
            print(f"\nRunning second variant: candle={variant2[0]}, macd={variant2[1]}")
            df2 = df.copy()
            res2 = run_analysis(df2, atype, window, variant2[0], variant2[1])
            export_markers_to_csv(df2, f"variant2_{variant2[0]}_{variant2[1]}.csv", res2, variant2[0], variant2[1])

        # Plot (Index auf 'date')
        df = df.set_index('date')
        if df2 is not None:
            df2 = df2.set_index('date')
        plot_with_plotly(df, analysis_results, counts, df2, res2)

    else:
        run_doe_analysis(df)
