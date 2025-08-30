# Mainframe_RTv250829_mplfinance.py
# Vollständige Version – NUR Ergänzungen/Korrekturen:
# - Legendengruppierung exakt (V1/V2 → Typ → Richtung, Classic/Hidden zusammen)
# - Zählerbox pro Variante
# - DOE: Standard = NUR 2×2-Facet-HTML mit allen 4 Typen (Zellen-Text = Classic + Hidden)
#        Optional per Flags: Gesamt-Heatmap & Einzelseiten

import os
import glob
import subprocess
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
import time

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# optionale Excel-Unterstützung
try:
    import openpyxl  # noqa: F401
    _HAS_XLSX = True
except Exception:
    _HAS_XLSX = False

# --- externe (bestehende) Module: NICHT ändern ---
from CBullDivg_Analysis_vectorized import CBullDivg_analysis
from CBullDivg_x2_analysis_vectorized import CBullDivg_x2_analysis
from HBearDivg_analysis_vectorized import HBearDivg_analysis
from HBullDivg_analysis_vectorized import HBullDivg_analysis
from Initialize_RSI_EMA_MACD import Initialize_RSI_EMA_MACD
from Local_Maximas_Minimas import Local_Max_Min
from Backtest_Divergences import (
    BacktestParams,
    prompt_params as backtest_prompt_params,
    backtest as run_backtest,
    export_backtest_xlsx,
)


# -------------------------------------------------
# Backtest-Zeitraum Auswahl (PowerShell Calendar mit Fallback)
# -------------------------------------------------
def select_backtest_timespan(df):
    """Return (start_dt, end_dt) for backtest. Uses Windows Forms calendar; fallback to console.
    Defaults to full span if cancelled/empty.
    """
    if 'date' not in df.columns:
        return None, None
    min_dt = pd.to_datetime(df['date']).min()
    max_dt = pd.to_datetime(df['date']).max()
    print(f"Available data range: {min_dt} to {max_dt}")

    # detect if intraday (has times or sub-daily interval)
    intraday = False
    try:
        times = pd.to_datetime(df['date'])
        if any(getattr(ts, 'hour', 0) != 0 or getattr(ts, 'minute', 0) != 0 for ts in times.head(100)):
            intraday = True
        else:
            dt_diffs = times.sort_values().diff().dropna()
            if not dt_diffs.empty and dt_diffs.min() < pd.Timedelta(days=1):
                intraday = True
    except Exception:
        pass

    # Try PowerShell date picker dialog
    try:
        ps = r'''
Add-Type -AssemblyName System.Windows.Forms
Add-Type -AssemblyName System.Drawing

$form = New-Object System.Windows.Forms.Form
$form.Text = "Select Backtest Date Range"
$form.Size = New-Object System.Drawing.Size(400,220)
$form.StartPosition = "CenterScreen"

$lbl1 = New-Object System.Windows.Forms.Label
$lbl1.Text = "Start date:"; $lbl1.Location = New-Object System.Drawing.Point(20,20); $lbl1.AutoSize = $true
$dp1 = New-Object System.Windows.Forms.DateTimePicker
$dp1.Format = [System.Windows.Forms.DateTimePickerFormat]::Short
$dp1.Location = New-Object System.Drawing.Point(120,16)

$lbl2 = New-Object System.Windows.Forms.Label
$lbl2.Text = "End date:"; $lbl2.Location = New-Object System.Drawing.Point(20,60); $lbl2.AutoSize = $true
$dp2 = New-Object System.Windows.Forms.DateTimePicker
$dp2.Format = [System.Windows.Forms.DateTimePickerFormat]::Short
$dp2.Location = New-Object System.Drawing.Point(120,56)

$ok = New-Object System.Windows.Forms.Button
$ok.Text = "OK"; $ok.Location = New-Object System.Drawing.Point(120,110)
$cancel = New-Object System.Windows.Forms.Button
$cancel.Text = "Cancel"; $cancel.Location = New-Object System.Drawing.Point(200,110)
$def = New-Object System.Windows.Forms.Button
$def.Text = "Default"; $def.Location = New-Object System.Drawing.Point(280,110)

$form.Controls.AddRange(@($lbl1,$dp1,$lbl2,$dp2,$ok,$cancel,$def))

$start=[datetime]::Parse("''' + str(pd.to_datetime(pd.to_datetime(min_dt).date())) + '''")
$end=[datetime]::Parse("''' + str(pd.to_datetime(pd.to_datetime(max_dt).date())) + '''")
$dp1.Value = $start; $dp2.Value = $end

$form.Tag = "DEFAULT"
$ok.Add_Click({ $form.Tag = "OK"; $form.Close() })
$cancel.Add_Click({ $form.Tag = "CANCEL"; $form.Close() })
$def.Add_Click({ $form.Tag = "DEFAULT"; $form.Close() })

$form.Topmost = $true
$form.Add_Shown({ $form.Activate() })
[void]$form.ShowDialog()

function ToUnix($dt){
  $dto = [System.DateTimeOffset]$dt
  return $dto.ToUniversalTime().ToUnixTimeSeconds()
}
$mode = $form.Tag
if ($mode -eq "OK") { Write-Output ("OK|{0}|{1}" -f (ToUnix $dp1.Value), (ToUnix $dp2.Value)) }
elseif ($mode -eq "DEFAULT") { Write-Output ("DEFAULT|{0}|{1}" -f (ToUnix $start), (ToUnix $end)) }
else { Write-Output ("DEFAULT|{0}|{1}" -f (ToUnix $start), (ToUnix $end)) }
'''
        # If intraday, switch to date+time format by injecting CustomFormat
        if intraday:
            ps = ps.replace('[System.Windows.Forms.DateTimePickerFormat]::Short', '[System.Windows.Forms.DateTimePickerFormat]::Custom')
            ps = ps.replace('$dp1.Location', '$dp1.CustomFormat = "yyyy-MM-dd HH:mm"; $dp1.ShowUpDown = $true; $dp1.Location')
            ps = ps.replace('$dp2.Location', '$dp2.CustomFormat = "yyyy-MM-dd HH:mm"; $dp2.ShowUpDown = $true; $dp2.Location')

        result = subprocess.run(["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps],
                                capture_output=True, text=True, timeout=60)
        if result.returncode == 0 and result.stdout:
            s = result.stdout.strip()
            if os.getenv('BT_DEBUG_TS'):
                print(f"[DEBUG] Date picker raw output: {s}")
            if "|" in s:
                parts = s.split("|")
                if len(parts) == 3:
                    mode, a, b = parts
                else:
                    mode, a, b = "", parts[0], parts[1]
                # parse epoch seconds robustly
                try:
                    s_dt = pd.to_datetime(int(float(a)), unit='s', utc=True)
                    e_dt = pd.to_datetime(int(float(b)), unit='s', utc=True)
                except Exception:
                    s_dt = pd.to_datetime(a, utc=True, errors='coerce')
                    e_dt = pd.to_datetime(b, utc=True, errors='coerce')
                # clamp to available range
                if pd.notna(s_dt) and s_dt < min_dt:
                    s_dt = pd.to_datetime(min_dt, utc=True)
                if pd.notna(e_dt) and e_dt > max_dt:
                    e_dt = pd.to_datetime(max_dt, utc=True)
                if pd.notna(s_dt) and pd.notna(e_dt):
                    # if mode explicitly DEFAULT, return full span
                    if str(mode).upper() == 'DEFAULT':
                        return pd.to_datetime(min_dt, utc=True), pd.to_datetime(max_dt, utc=True)
                    if os.getenv('BT_DEBUG_TS'):
                        print(f"[DEBUG] Parsed picker: mode={mode}, start={s_dt}, end={e_dt}")
                    return s_dt, e_dt
    except Exception:
        pass

    # If dialog failed or user canceled without output -> use full span by default
    return pd.to_datetime(min_dt, utc=True), pd.to_datetime(max_dt, utc=True)


# -------------------------------------------------
# Datei-Auswahl (PowerShell Dialog mit Fallback)
# -------------------------------------------------
# Standard-Default-Datei für Input
DEFAULT_INPUT_FILE = r"C:\Projekte\crt_250816\data\raw\btc_1day_candlesticks_all.csv"


def get_input_file():
    """PowerShell-OpenFileDialog; Fallback: Konsole."""
    try:
        ps_command = r'''
Add-Type -AssemblyName System.Windows.Forms
$dlg = New-Object System.Windows.Forms.OpenFileDialog
$dlg.Filter = "CSV/Parquet (*.csv;*.parquet)|*.csv;*.parquet|CSV (*.csv)|*.csv|Parquet (*.parquet)|*.parquet|All files (*.*)|*.*"
$dlg.Title = "Select CSV/Parquet input file"
$dlg.InitialDirectory = "C:\Projekte\crt_250816\data\raw"
if ($dlg.ShowDialog() -eq [System.Windows.Forms.DialogResult]::OK) { Write-Output $dlg.FileName }
'''
        result = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_command],
            capture_output=True, text=True, timeout=60
        )
        path = (result.stdout or "").strip()
        if result.returncode == 0 and path:
            return path
        # Nichts gewählt -> Standard verwenden
        return DEFAULT_INPUT_FILE
    except Exception:
        return DEFAULT_INPUT_FILE


def get_input_file_console():
    default_file = DEFAULT_INPUT_FILE
    print(f"\nDefault file: {default_file}")
    print("Press Enter to use default, or choose from available files:")

    candidates = []
    search_patterns = [
        "*.csv", "*.parquet",
        "../*/*.csv", "../*/*.parquet",
        "../../data/raw/*.csv", "../../data/raw/*.parquet",
        "../../data/*.csv", "../../data/*.parquet"
    ]
    for pat in search_patterns:
        candidates.extend(glob.glob(pat))

    for i, f in enumerate(candidates, 1):
        print(f"{i}: {f}")
    print(f"{len(candidates) + 1}: Enter custom path")
    print(f"0 or Enter: Use default ({default_file})")

    while True:
        try:
            choice = input(f"\nSelect file (0-{len(candidates) + 1}): ").strip()
            if choice in ("", "0"):
                return default_file
            if choice.isdigit():
                idx = int(choice)
                if 1 <= idx <= len(candidates):
                    return candidates[idx - 1]
                if idx == len(candidates) + 1:
                    custom = input("Enter full path (or Enter for default): ").strip()
                    return custom if custom else default_file
            print("Invalid choice. Try again.")
        except (EOFError, KeyboardInterrupt):
            return default_file


# -------------------------------------------------
# Parameter/Typ-Auswahl
# -------------------------------------------------
def get_analysis_parameters():
    import sys
    # CLI: script <type> <candle> <macd>
    if len(sys.argv) > 2:
        try:
            candle_percent = float(sys.argv[2])
            macd_percent = float(sys.argv[3]) if len(sys.argv) > 3 else 3.25
            return candle_percent, macd_percent, None
        except Exception:
            pass

    try:
        print("\nEnter analysis parameters (or press Enter for defaults):")
        s = input("Enter Candle tolerance % (default: 0.1): ").strip()
        candle_percent = 0.1 if s == "" else float(s)
        s = input("Enter MACD tolerance % (default: 3.25): ").strip()
        macd_percent = 3.25 if s == "" else float(s)

        print("\nOptional second variant for comparison:")
        s1 = input("Enter second Candle tolerance % (leave empty for none): ").strip()
        s2 = input("Enter second MACD tolerance % (leave empty for none): ").strip()

        variant2 = None
        if s1 or s2:
            try:
                c2 = float(s1) if s1 else candle_percent
                m2 = float(s2) if s2 else macd_percent
                variant2 = (c2, m2)
            except Exception as e:
                print(f"Invalid second variant: {e}")
                variant2 = None

        return candle_percent, macd_percent, variant2
    except (EOFError, KeyboardInterrupt):
        return 0.1, 3.25, None


def get_analysis_type():
    import sys
    if len(sys.argv) > 1:
        c = (sys.argv[1] or "").lower().strip()
        if c in list("abcdefg"):
            return c

    try:
        print("\nSelect analysis type:")
        print("a: CBullDivg_Analysis (Classic Bullish Divergence)")
        print("b: BullDivg_x2_analysis (Extended Bullish Divergence)")
        print("c: HBearDivg_analysis (Hidden Bearish Divergence)")
        print("d: HBullDivg_analysis (Hidden Bullish Divergence)")
        print("e: All analyses (a-d)")
        print("f: DOE (Design of Experiments)")
        print("g: Backtest markers (validate signals)")
        print("h: Backtest DOE markers (combine doe_markers_*.csv)")
        while True:
            c = input("\nEnter your choice (a-h): ").lower().strip()
            if c in list("abcdefgh"):
                return c
            print("Invalid choice. Please enter a, b, c, d, e, f, g, or h.")
    except (EOFError, KeyboardInterrupt):
        return "e"


# -------------------------------------------------
# Analysen laufen lassen (ruft vorhandene Module)
# -------------------------------------------------
def run_analysis(df, analysis_type, window=5, candle_tol=0.1, macd_tol=3.25):
    results = {}
    if analysis_type in ['a', 'e', 'f']:
        CBullDivg_analysis(df, window, candle_tol, macd_tol)
        results['CBullDivg'] = True
    if analysis_type in ['b', 'e', 'f']:
        CBullDivg_x2_analysis(df, window, candle_tol, macd_tol)
        results['CBullDivg_x2'] = True
    if analysis_type in ['c', 'e', 'f']:
        HBearDivg_analysis(df, window, candle_tol, macd_tol)
        results['HBearDivg'] = True
    if analysis_type in ['d', 'e', 'f']:
        HBullDivg_analysis(df, window, candle_tol, macd_tol)
        results['HBullDivg'] = True
    return results


# -------------------------------------------------
# Marker-Export (CSV/XLSX mit Summary)
# -------------------------------------------------
def export_markers_to_csv(df, filename, analysis_results, candle_percent, macd_percent):
    os.makedirs('results', exist_ok=True)
    markers = []
    counts = {k: {'classic': 0, 'hidden': 0, 'total': 0}
              for k, v in analysis_results.items() if v}

    def _append(t, dt):
        markers.append({
            'Type': t,
            'Date': dt,
            'Candle_Percent': candle_percent,
            'MACD_Percent': macd_percent
        })

    for i in range(len(df)):
        dt = df['date'].iloc[i]
        # Classic Bullish
        if analysis_results.get('CBullDivg', False):
            if 'CBullD_gen' in df.columns and pd.notna(df.at[i, 'CBullD_gen']) and df.at[i, 'CBullD_gen'] == 1:
                _append('CBullDivg_Classic', dt)
                counts['CBullDivg']['classic'] += 1
            if 'CBullD_neg_MACD' in df.columns and pd.notna(df.at[i, 'CBullD_neg_MACD']) and df.at[i, 'CBullD_neg_MACD'] == 1:
                _append('CBullDivg_Hidden', dt)
                counts['CBullDivg']['hidden'] += 1

        # x2
        if analysis_results.get('CBullDivg_x2', False):
            if 'CBullD_x2_gen' in df.columns and pd.notna(df.at[i, 'CBullD_x2_gen']) and df.at[i, 'CBullD_x2_gen'] == 1:
                _append('CBullDivg_x2_Classic', dt)
                counts['CBullDivg_x2']['classic'] += 1

        # Hidden Bearish
        if analysis_results.get('HBearDivg', False):
            if 'HBearD_gen' in df.columns and pd.notna(df.at[i, 'HBearD_gen']) and df.at[i, 'HBearD_gen'] == 1:
                _append('HBearDivg_Classic', dt)
                counts['HBearDivg']['classic'] += 1

        # Hidden Bullish
        if analysis_results.get('HBullDivg', False):
            if 'HBullD_gen' in df.columns and pd.notna(df.at[i, 'HBullD_gen']) and df.at[i, 'HBullD_gen'] == 1:
                _append('HBullDivg_Classic', dt)
                counts['HBullDivg']['classic'] += 1
            if 'HBullD_neg_MACD' in df.columns and pd.notna(df.at[i, 'HBullD_neg_MACD']) and df.at[i, 'HBullD_neg_MACD'] == 1:
                _append('HBullDivg_Hidden', dt)
                counts['HBullDivg']['hidden'] += 1

    # Totale
    for k in counts:
        counts[k]['total'] = counts[k]['classic'] + counts[k]['hidden']

    print("\n=== ANALYSIS SUMMARY ===")
    for name, c in counts.items():
        print(f"{name}: Classic={c['classic']}, Hidden={c['hidden']}, Total={c['total']}")
    print(f"\nGrand Total markers found: {len(markers)}")

    # CSV
    csv_path = os.path.join('results', filename)
    pd.DataFrame(markers).to_csv(csv_path, index=False)
    print(f"\nMarkers exported to {csv_path}")

    # XLSX (optional)
    if _HAS_XLSX:
        xlsx_path = csv_path.replace('.csv', '.xlsx')
        df_x = pd.DataFrame(markers).copy()
        if not df_x.empty:
            df_x['Date'] = pd.to_datetime(df_x['Date']).dt.tz_localize(None)
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            # alle Marker (sortiert)
            df_x.sort_values(['Type', 'Date']).to_excel(writer, sheet_name='All_Markers', index=False)
            # pro Typ eigenes Blatt
            for t in df_x['Type'].dropna().unique():
                df_x[df_x['Type'] == t].sort_values('Date').to_excel(
                    writer, sheet_name=str(t).replace('_', ' ')[:31], index=False
                )
            # Summary
            s = (pd.DataFrame.from_dict(counts, orient='index')[['classic', 'hidden', 'total']]
                 .rename_axis('Analysis').reset_index())
            s.to_excel(writer, sheet_name='Summary', index=False)
            ws = writer.sheets['Summary']
            grand_row = len(s) + 2
            ws.cell(row=grand_row, column=1).value = 'Grand Total'
            ws.cell(row=grand_row, column=2).value = f'=SUM(B2:B{grand_row - 1})'
            ws.cell(row=grand_row, column=3).value = f'=SUM(C2:C{grand_row - 1})'
            ws.cell(row=grand_row, column=4).value = f'=SUM(D2:D{grand_row - 1})'
        print(f"Enhanced XLSX exported to {xlsx_path}")

    return counts


# -------------------------------------------------
# Plot-Hilfen
# -------------------------------------------------
def _safe_get_vals(df_idx, dt):
    """Liefert Low, High, RSI, MACD für gegebenes Datum; NaN falls nicht vorhanden."""
    try:
        row = df_idx.loc[dt]
        return row.get('low', np.nan), row.get('high', np.nan), row.get('RSI', np.nan), row.get('macd_histogram', np.nan)
    except Exception:
        return np.nan, np.nan, np.nan, np.nan


def add_markers_to_plotly(fig, df, analysis_results, variant_name):
    """
    Zeichnet Marker in 3 Subplots (Price/RSI/MACD).
    Legendengruppierung wie gefordert (Classic/Hidden gemeinsam).
    Rückgabe: Set aus (date, y_price) zur Varianten-Differenz.
    """
    positions = set()

    marker_styles = {
        'CBullDivg':    {'symbol': 'triangle-up',   'color': 'green'},
        'CBullDivg_x2': {'symbol': 'triangle-down', 'color': 'green'},
        'HBullDivg':    {'symbol': 'diamond',       'color': 'green'},
        'HBearDivg':    {'symbol': 'square',        'color': 'red'}
    }

    def _group_id(div_type, direction):
        return f"{variant_name}|{div_type}|{direction}"

    def _proxy_name(div_type, direction):
        return f"{variant_name} - {div_type} - {direction}"

    def _add_proxy_legend_entry(div_type, direction):
        gid = _group_id(div_type, direction)
        fig.add_trace(
            go.Scatter(
                x=[df['date'].iloc[0] if len(df) else None],
                y=[np.nan],
                mode='markers',
                name=_proxy_name(div_type, direction),
                marker=dict(
                    symbol=marker_styles[div_type]['symbol'],
                    size=12,
                    line=dict(width=2),
                ),
                visible='legendonly',
                legendgroup=gid,
                legendgrouptitle_text=variant_name,
                showlegend=True,
                hoverinfo='skip'
            ),
            row=1, col=1
        )

    # Proxy-Gruppen anlegen (Sichtbarkeit in Legende sichern)
    if analysis_results.get('CBullDivg', False):
        _add_proxy_legend_entry('CBullDivg', 'Bullish')
    if analysis_results.get('CBullDivg_x2', False):
        _add_proxy_legend_entry('CBullDivg_x2', 'Bullish')
    if analysis_results.get('HBullDivg', False):
        _add_proxy_legend_entry('HBullDivg', 'Bullish')
    if analysis_results.get('HBearDivg', False):
        _add_proxy_legend_entry('HBearDivg', 'Bearish')

    def _add_pair(dates, y_price, y_rsi, y_macd, div_type, direction, size, opacity):
        gid = _group_id(div_type, direction)
        common_marker = dict(symbol=marker_styles[div_type]['symbol'],
                             color=marker_styles[div_type]['color'],
                             size=size, opacity=opacity, line=dict(width=2))
        # Preis
        fig.add_trace(
            go.Scatter(x=dates, y=y_price, mode='markers',
                       name=f"{div_type} - {direction}",
                       marker=common_marker,
                       legendgroup=gid,
                       showlegend=False),
            row=1, col=1
        )
        # RSI
        fig.add_trace(
            go.Scatter(x=dates, y=y_rsi, mode='markers',
                       name=f"{div_type} - {direction}",
                       marker=common_marker,
                       legendgroup=gid,
                       showlegend=False),
            row=2, col=1
        )
        # MACD
        fig.add_trace(
            go.Scatter(x=dates, y=y_macd, mode='markers',
                       name=f"{div_type} - {direction}",
                       marker=common_marker,
                       legendgroup=gid,
                       showlegend=False),
            row=3, col=1
        )

    df_idx = df.set_index('date')

    for i in range(len(df)):
        # Classic Bullish
        if analysis_results.get('CBullDivg', False) and 'CBullD_gen' in df.columns and df.at[i, 'CBullD_gen'] == 1:
            d1 = pd.to_datetime(df.at[i, 'CBullD_Lower_Low_date_gen'])
            d2 = pd.to_datetime(df.at[i, 'CBullD_Higher_Low_date_gen'])
            low1, _, rsi1, macd1 = _safe_get_vals(df_idx, d1)
            low2, _, rsi2, macd2 = _safe_get_vals(df_idx, d2)
            if pd.notna(low1) and pd.notna(low2):
                _add_pair([d1, d2], [low1 * 0.99, low2 * 0.99], [rsi1, rsi2], [macd1, macd2],
                          'CBullDivg', 'Bullish', size=12, opacity=1.0)
                positions.update([(d1, low1 * 0.99), (d2, low2 * 0.99)])

        # Hidden Bullish (neg MACD)
        if analysis_results.get('CBullDivg', False) and 'CBullD_neg_MACD' in df.columns and df.at[i, 'CBullD_neg_MACD'] == 1:
            d1 = pd.to_datetime(df.at[i, 'CBullD_Lower_Low_date_neg_MACD'])
            d2 = pd.to_datetime(df.at[i, 'CBullD_Higher_Low_date_neg_MACD'])
            low1, _, rsi1, macd1 = _safe_get_vals(df_idx, d1)
            low2, _, rsi2, macd2 = _safe_get_vals(df_idx, d2)
            if pd.notna(low1) and pd.notna(low2):
                _add_pair([d1, d2], [low1 * 0.98, low2 * 0.98], [rsi1, rsi2], [macd1, macd2],
                          'CBullDivg', 'Bullish', size=10, opacity=0.7)
                positions.update([(d1, low1 * 0.98), (d2, low2 * 0.98)])

        # x2 Bullish
        if analysis_results.get('CBullDivg_x2', False) and 'CBullD_x2_gen' in df.columns and df.at[i, 'CBullD_x2_gen'] == 1:
            d1 = pd.to_datetime(df.at[i, 'CBullD_x2_Lower_Low_date_gen'])
            d2 = pd.to_datetime(df.at[i, 'CBullD_x2_Higher_Low_date_gen'])
            low1, _, rsi1, macd1 = _safe_get_vals(df_idx, d1)
            low2, _, rsi2, macd2 = _safe_get_vals(df_idx, d2)
            if pd.notna(low1) and pd.notna(low2):
                _add_pair([d1, d2], [low1 * 0.97, low2 * 0.97], [rsi1, rsi2], [macd1, macd2],
                          'CBullDivg_x2', 'Bullish', size=12, opacity=1.0)
                positions.update([(d1, low1 * 0.97), (d2, low2 * 0.97)])

        # Hidden Bearish
        if analysis_results.get('HBearDivg', False) and 'HBearD_gen' in df.columns and df.at[i, 'HBearD_gen'] == 1:
            d1 = pd.to_datetime(df.at[i, 'HBearD_Higher_High_date_gen'])
            d2 = pd.to_datetime(df.at[i, 'HBearD_Lower_High_date_gen'])
            _, high1, rsi1, macd1 = _safe_get_vals(df_idx, d1)
            _, high2, rsi2, macd2 = _safe_get_vals(df_idx, d2)
            if pd.notna(high1) and pd.notna(high2):
                _add_pair([d1, d2], [high1 * 1.01, high2 * 1.01], [rsi1, rsi2], [macd1, macd2],
                          'HBearDivg', 'Bearish', size=12, opacity=1.0)
                positions.update([(d1, high1 * 1.01), (d2, high2 * 1.01)])

        # Hidden Bullish (gen)
        if analysis_results.get('HBullDivg', False) and 'HBullD_gen' in df.columns and df.at[i, 'HBullD_gen'] == 1:
            d1 = pd.to_datetime(df.at[i, 'HBullD_Lower_Low_date_gen'])
            d2 = pd.to_datetime(df.at[i, 'HBullD_Higher_Low_date_gen'])
            low1, _, rsi1, macd1 = _safe_get_vals(df_idx, d1)
            low2, _, rsi2, macd2 = _safe_get_vals(df_idx, d2)
            if pd.notna(low1) and pd.notna(low2):
                _add_pair([d1, d2], [low1 * 0.96, low2 * 0.96], [rsi1, rsi2], [macd1, macd2],
                          'HBullDivg', 'Bullish', size=12, opacity=1.0)
                positions.update([(d1, low1 * 0.96), (d2, low2 * 0.96)])

        # Hidden Bullish (neg MACD)
        if analysis_results.get('HBullDivg', False) and 'HBullD_neg_MACD' in df.columns and df.at[i, 'HBullD_neg_MACD'] == 1:
            d1 = pd.to_datetime(df.at[i, 'HBullD_Lower_Low_date_neg_MACD'])
            d2 = pd.to_datetime(df.at[i, 'HBullD_Higher_Low_date_neg_MACD'])
            low1, _, rsi1, macd1 = _safe_get_vals(df_idx, d1)
            low2, _, rsi2, macd2 = _safe_get_vals(df_idx, d2)
            if pd.notna(low1) and pd.notna(low2):
                _add_pair([d1, d2], [low1 * 0.95, low2 * 0.95], [rsi1, rsi2], [macd1, macd2],
                          'HBullDivg', 'Bullish', size=10, opacity=0.7)
                positions.update([(d1, low1 * 0.95), (d2, low2 * 0.95)])

    return positions


def _format_counts_box(title, counts):
    """Erzeugt HTML-Text für eine Variante (Titel + per-Analysis-Zeilen + Variant Total)."""
    if not counts:
        return ""
    lines = [f"<b>{title}</b>"]
    v_total = 0
    for name, d in counts.items():
        lines.append(f"{name}: C={d['classic']}, H={d['hidden']}, T={d['total']}")
        v_total += d['total']
    lines.append(f"<b>{title} Total: {v_total}</b>")
    return "<br>".join(lines)


def plot_with_plotly(df_main, analysis_results, counts_main, df_var2=None, variant2_results=None, counts_var2=None, asset_label=None):
    fig = make_subplots(
        rows=3, cols=1, shared_xaxes=True, vertical_spacing=0.05,
        subplot_titles=('Price', 'RSI', 'MACD Histogram'),
        row_heights=[0.4, 0.3, 0.3]
    )

    # Candles
    fig.add_trace(
        go.Candlestick(
            x=df_main.index, open=df_main['open'], high=df_main['high'],
            low=df_main['low'], close=df_main['close'], name='Price',
            increasing_line_color='#44ff44', decreasing_line_color='#ff4444'
        ),
        row=1, col=1
    )

    # EMAs (vorausgesetzt von Initialize_RSI_EMA_MACD)
    ema_defs = {'EMA_20': 'yellow', 'EMA_50': 'cyan', 'EMA_100': 'magenta', 'EMA_200': 'orange'}
    for col, _col_color in ema_defs.items():
        if col in df_main.columns:
            fig.add_trace(
                go.Scatter(x=df_main.index, y=df_main[col], mode='lines', name=col,
                           line=dict(width=1), opacity=0.8),
                row=1, col=1
            )

    # RSI
    if 'RSI' in df_main.columns:
        fig.add_trace(
            go.Scatter(x=df_main.index, y=df_main['RSI'], mode='lines', name='RSI', line=dict(width=2)),
            row=2, col=1
        )

    # MACD-Histogramm
    macd_vals = df_main['macd_histogram'] if 'macd_histogram' in df_main.columns else pd.Series(0, index=df_main.index)
    colors = ['#00FF00' if v >= 0 else '#FF0000' for v in macd_vals]
    fig.add_trace(
        go.Bar(x=df_main.index, y=macd_vals, name='MACD Hist.', marker_color=colors),
        row=3, col=1
    )

    # Marker (Variante 1)
    df_m_for_markers = df_main.reset_index().rename(columns={'index': 'date'})
    main_positions = add_markers_to_plotly(fig, df_m_for_markers, analysis_results, 'V1 (Main)')

    # Marker (Variante 2) + Vergleich Outlines
    if df_var2 is not None and variant2_results is not None:
        df_v2_for_markers = df_var2.reset_index().rename(columns={'index': 'date'})
        v2_positions = add_markers_to_plotly(fig, df_v2_for_markers, variant2_results, 'V2 (Variant)')

        additional = v2_positions - main_positions
        missing = main_positions - v2_positions

        def _add_outline(dates_prices, name, color):
            if not dates_prices:
                return
            dates, prices = zip(*dates_prices)
            rsi_vals = [df_m_for_markers.set_index('date').get('RSI', pd.Series(dtype=float)).get(d, np.nan) for d in dates]
            macd_vals_local = [df_m_for_markers.set_index('date').get('macd_histogram', pd.Series(dtype=float)).get(d, np.nan) for d in dates]

            mk = dict(symbol='circle-open', color=color, size=18, line=dict(width=1.5))
            # Preis
            fig.add_trace(go.Scatter(x=list(dates), y=list(prices), mode='markers',
                                     name=name, marker=mk,
                                     legendgroup=f"Comparison|{name}",
                                     legendgrouptitle_text="Comparison",
                                     showlegend=True),
                          row=1, col=1)
            # RSI
            fig.add_trace(go.Scatter(x=list(dates), y=rsi_vals, mode='markers',
                                     name=name, marker=mk,
                                     legendgroup=f"Comparison|{name}",
                                     showlegend=False),
                          row=2, col=1)
            # MACD
            fig.add_trace(go.Scatter(x=list(dates), y=macd_vals_local, mode='markers',
                                     name=name, marker=mk,
                                     legendgroup=f"Comparison|{name}",
                                     showlegend=False),
                          row=3, col=1)

        _add_outline(additional, 'V2 Additional', 'yellow')
        _add_outline(missing, 'V2 Missing', 'blue')

    # Layout/Legend
    title_txt = 'Technical Analysis with Divergence Markers'
    if asset_label:
        title_txt = f"{asset_label} – {title_txt}"
    fig.update_layout(
        title_text=title_txt,
        xaxis_rangeslider_visible=False,
        template='plotly_dark',
        legend=dict(
            traceorder='grouped',
            groupclick='togglegroup',  # Gruppe (Classic+Hidden) gemeinsam toggeln
            itemclick='toggle'
        ),
        height=1200
    )
    fig.update_yaxes(title_text="Price", row=1, col=1)
    fig.update_yaxes(title_text="RSI", range=[0, 100], row=2, col=1)
    fig.update_yaxes(title_text="MACD", row=3, col=1)

    # ======= Zählerbox (für alle vorhandenen Varianten) =======
    box_parts = []
    box_parts.append(_format_counts_box("V1 (Main)", counts_main))
    if counts_var2:
        box_parts.append(_format_counts_box("V2 (Variant)", counts_var2))
    box_text = "<br><br>".join([p for p in box_parts if p])

    if box_text:
        fig.add_annotation(text=box_text, align='left', showarrow=False,
                           xref='paper', yref='paper', x=0.01, y=0.99,
                           bgcolor="rgba(0,0,0,0.7)", bordercolor="white", borderwidth=1)

    # speichern & anzeigen
    os.makedirs('results', exist_ok=True)
    out_html = os.path.join('results', 'plot.html')
    fig.write_html(out_html)
    print(f"\nInteractive plot saved to: {out_html}")
    fig.show()


# -------------------------------------------------
# DOE – Visualisierungen
# -------------------------------------------------
def _doe_plot_total(df_counts, asset_label=None):
    """Gesamt-Heatmap (Total über alle Analysen)."""
    agg = (df_counts
           .groupby(['Candle_Percent', 'MACD_Percent'], as_index=False)['Total']
           .sum())
    if agg.empty:
        print("No DOE data for total heatmap.")
        return
    pivot = agg.pivot(index='Candle_Percent', columns='MACD_Percent', values='Total').fillna(0)

    hm = go.Figure(data=go.Heatmap(
        z=pivot.values,
        x=pivot.columns.tolist(),
        y=pivot.index.tolist(),
        colorbar=dict(title='Total markers')
    ))
    title = "DOE – Total markers (sum over all analyses)"
    if asset_label:
        title = f"{asset_label} – {title}"
    hm.update_layout(
        title=title,
        xaxis_title="MACD tolerance (%)",
        yaxis_title="Candle tolerance (%)",
        template="plotly_dark",
        height=700
    )

    out_html = os.path.join('results', 'doe_heatmap_total.html')
    hm.write_html(out_html)
    print(f"DOE heatmap saved to: {out_html}")
    hm.show()


def _doe_plot_by_type(df_counts, single_pages=False, asset_label=None, overlay_points=None):
    """
    2×2-Facet mit einheitlicher Skala; Zellen-Text = Classic & Hidden.
    Optional: single_pages=True erzeugt je Typ eigene HTMLs zusätzlich.
    Erwartet Spalten: Candle_Percent, MACD_Percent, Analysis, Classic, Hidden, Total
    """
    analyses = ["CBullDivg", "CBullDivg_x2", "HBullDivg", "HBearDivg"]
    df_counts = df_counts.copy()
    if df_counts.empty:
        print("No DOE data for by-type heatmaps.")
        return

    # globale Z-Max für konsistente Farbschale
    zmax = (df_counts.groupby(['Candle_Percent', 'MACD_Percent', 'Analysis'])['Total']
            .sum().max())
    if pd.isna(zmax) or zmax <= 0:
        zmax = None  # Plotly wählt automatisch

    # Einzelseiten je Typ (optional)
    if single_pages:
        for a in analyses:
            grp = (df_counts[df_counts['Analysis'] == a]
                   .groupby(['Candle_Percent', 'MACD_Percent'], as_index=False)
                   .sum(numeric_only=True))
            if grp.empty:
                print(f"DOE: no data for {a}, skipping single heatmap.")
                continue

            pv_total   = grp.pivot(index='Candle_Percent', columns='MACD_Percent', values='Total').sort_index().sort_index(axis=1).fillna(0)
            pv_classic = grp.pivot(index='Candle_Percent', columns='MACD_Percent', values='Classic').reindex_like(pv_total).fillna(0)
            pv_hidden  = grp.pivot(index='Candle_Percent', columns='MACD_Percent', values='Hidden').reindex_like(pv_total).fillna(0)

            arr_c = pv_classic.astype(int).values
            arr_h = pv_hidden.astype(int).values
            text = np.char.add("C:", arr_c.astype(str))
            text = np.char.add(text, "\nH:")
            text = np.char.add(text, arr_h.astype(str))

            fig = go.Figure(data=go.Heatmap(
                z=pv_total.values, x=pv_total.columns.tolist(), y=pv_total.index.tolist(),
                zmin=0 if zmax is not None else None, zmax=zmax,
                text=text, texttemplate="%{text}",
                hovertemplate="Candle: %{y}<br>MACD: %{x}<br>Total: %{z}<br>%{text}<extra></extra>",
                colorbar=dict(title='Total')
            ))
            title = f"DOE – {a}"
            if asset_label:
                title = f"{asset_label} – {title}"
            fig.update_layout(
                title=title,
                xaxis_title="MACD tolerance (%)",
                yaxis_title="Candle tolerance (%)",
                template="plotly_dark",
                height=650
            )
            out_html = os.path.join('results', f"doe_heatmap_{a}.html")
            fig.write_html(out_html)
            print(f"DOE by-type heatmap saved to: {out_html}")
            fig.show()

    # Facet 2×2 (nur dieses HTML ist per Default aktiv)
    fig = make_subplots(rows=2, cols=2,
                        subplot_titles=analyses,
                        vertical_spacing=0.12, horizontal_spacing=0.08)
    rowcol = [(1,1),(1,2),(2,1),(2,2)]
    for (a, (r, c)) in zip(analyses, rowcol):
        grp = (df_counts[df_counts['Analysis'] == a]
               .groupby(['Candle_Percent', 'MACD_Percent'], as_index=False)
               .sum(numeric_only=True))
        if grp.empty:
            fig.add_annotation(text=f"No data for {a}", showarrow=False, row=r, col=c)
            continue

        pv_total   = grp.pivot(index='Candle_Percent', columns='MACD_Percent', values='Total').sort_index().sort_index(axis=1).fillna(0)
        pv_classic = grp.pivot(index='Candle_Percent', columns='MACD_Percent', values='Classic').reindex_like(pv_total).fillna(0)
        pv_hidden  = grp.pivot(index='Candle_Percent', columns='MACD_Percent', values='Hidden').reindex_like(pv_total).fillna(0)

        arr_c = pv_classic.astype(int).values
        arr_h = pv_hidden.astype(int).values
        text = np.char.add("C:", arr_c.astype(str))
        text = np.char.add(text, "\nH:")
        text = np.char.add(text, arr_h.astype(str))

        show_scale = (r, c) == (2, 2)  # Farbleiste nur im letzten Panel
        trace = go.Heatmap(
            z=pv_total.values, x=pv_total.columns.tolist(), y=pv_total.index.tolist(),
            zmin=0 if zmax is not None else None, zmax=zmax,
            text=text, texttemplate="%{text}",
            hovertemplate="Candle: %{y}<br>MACD: %{x}<br>Total: %{z}<br>%{text}<extra></extra>",
            showscale=show_scale, colorbar=dict(title='Total') if show_scale else None
        )
        fig.add_trace(trace, row=r, col=c)
        # Overlay robust top/Pareto points if provided
        if overlay_points and a in overlay_points:
            pts = overlay_points[a]
            if pts:
                xs = [p['MACD_Percent'] for p in pts]
                ys = [p['Candle_Percent'] for p in pts]
                labels = [p.get('label', '') for p in pts]
                fig.add_trace(
                    go.Scatter(x=xs, y=ys, mode='markers+text', text=labels,
                               textposition='top center',
                               marker=dict(symbol='x', size=12, color='white', line=dict(width=1)),
                               name=f"Top/Pareto – {a}"),
                    row=r, col=c
                )

    title = "DOE – markers by divergence type (C: Classic, H: Hidden)"
    if asset_label:
        title = f"{asset_label} – {title}"
    fig.update_layout(
        title=title,
        template="plotly_dark",
        height=950
    )
    # Achsentitel für alle Subplots
    for r in (1,2):
        for c in (1,2):
            fig.update_xaxes(title_text="MACD tolerance (%)", row=r, col=c)
            fig.update_yaxes(title_text="Candle tolerance (%)", row=r, col=c)

    out_html = os.path.join('results', 'doe_heatmaps_by_type.html')
    fig.write_html(out_html)
    print(f"DOE facet heatmaps saved to: {out_html}")
    fig.show()


# -------------------------------------------------
# DOE (inkl. Plots)
# -------------------------------------------------
def _doe_worker(idx, candle_tol, macd_tol, df, window):
    """
    Worker for DOE: clones df, runs analysis, exports markers CSV, returns counts and filename.
    Keeps behavior identical to sequential path by using export_markers_to_csv.
    """
    try:
        df_copy = df.copy()
        res = run_analysis(df_copy, 'e', window, candle_tol, macd_tol)
        out_csv = f"doe_markers_candle{candle_tol}_macd{macd_tol}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        counts = export_markers_to_csv(df_copy, out_csv, res, candle_tol, macd_tol)
        return {
            'idx': idx,
            'candle_tol': candle_tol,
            'macd_tol': macd_tol,
            'counts': counts,
            'out_csv': out_csv,
            'error': None,
        }
    except Exception as e:
        return {
            'idx': idx,
            'candle_tol': candle_tol,
            'macd_tol': macd_tol,
            'counts': None,
            'out_csv': None,
            'error': str(e),
        }


def run_doe_analysis(df, doe_params_file="doe_parameters_example.csv",
                     make_total=False, make_single=False, parallel=True, asset_label=None):
    """
    Standard: nur 2×2-Facet-HTML (alle 4 Typen) mit Classic/Hidden-Text.
    Optional: make_total=True → Gesamt-Heatmap zusätzlich,
              make_single=True → einzelne HTMLs je Typ zusätzlich.
    """
    print(f"Loading DOE parameters from {doe_params_file}...")
    try:
        prm = pd.read_csv(doe_params_file)
        required = {'candle_percent', 'macd_percent'}
        if not required.issubset(set(prm.columns)):
            raise ValueError(f"DOE parameters file must contain columns: {sorted(required)}")
    except Exception as e:
        print(f"Error loading DOE parameters: {e}. Exiting DOE analysis.")
        return

    window = 5
    all_counts_rows = []
    all_marker_frames = []

    # Maintain original processing order deterministically
    indexed_params = [(idx, float(row['candle_percent']), float(row['macd_percent'])) for idx, row in prm.iterrows()]
    t0 = time.perf_counter()
    if not parallel or len(indexed_params) <= 1:
        # Original sequential behavior
        for idx, candle_tol, macd_tol in indexed_params:
            print(f"\nRunning DOE iteration {idx + 1}: candle_tol={candle_tol}, macd_tol={macd_tol}")

            df_copy = df.copy()
            res = run_analysis(df_copy, 'e', window, candle_tol, macd_tol)

            out_csv = f"doe_markers_candle{candle_tol}_macd{macd_tol}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            counts = export_markers_to_csv(df_copy, out_csv, res, candle_tol, macd_tol)

            for a_name, c_dict in counts.items():
                all_counts_rows.append({
                    'Candle_Percent': candle_tol,
                    'MACD_Percent': macd_tol,
                    'Analysis': a_name,
                    'Classic': c_dict['classic'],
                    'Hidden': c_dict['hidden'],
                    'Total': c_dict['total']
                })

            try:
                df_m = pd.read_csv(os.path.join('results', out_csv))
                df_m['Candle_Percent'] = candle_tol
                df_m['MACD_Percent'] = macd_tol
                all_marker_frames.append(df_m)
            except Exception:
                pass
    else:
        # Parallel execution using ProcessPool for CPU-bound DOE
        max_workers_env = os.getenv('DOE_MAX_WORKERS')
        try:
            max_workers = int(max_workers_env) if max_workers_env else min(len(indexed_params), (os.cpu_count() or 4))
        except Exception:
            max_workers = min(len(indexed_params), (os.cpu_count() or 4))

        print(f"Running DOE in parallel (ProcessPool) with {max_workers} workers...")
        results = []
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(_doe_worker, idx, c, m, df, window) for idx, c, m in indexed_params]
            for f in as_completed(futures):
                try:
                    results.append(f.result())
                except Exception as e:
                    results.append({'idx': -1, 'error': str(e)})

        # Sort back to original DataFrame order of DOE params
        results.sort(key=lambda r: r['idx'])

        # Aggregate exactly like sequential path
        for r in results:
            if r.get('error'):
                print(f"DOE iteration {r['idx'] + 1} failed: {r['error']}")
                continue
            candle_tol = r['candle_tol']
            macd_tol = r['macd_tol']
            counts = r['counts'] or {}
            out_csv = r['out_csv']

            for a_name, c_dict in counts.items():
                all_counts_rows.append({
                    'Candle_Percent': candle_tol,
                    'MACD_Percent': macd_tol,
                    'Analysis': a_name,
                    'Classic': c_dict['classic'],
                    'Hidden': c_dict['hidden'],
                    'Total': c_dict['total']
                })

            try:
                df_m = pd.read_csv(os.path.join('results', out_csv))
                df_m['Candle_Percent'] = candle_tol
                df_m['MACD_Percent'] = macd_tol
                all_marker_frames.append(df_m)
            except Exception:
                pass

    elapsed = time.perf_counter() - t0
    print(f"DOE completed in {elapsed:.2f}s across {len(indexed_params)} parameter pairs.")

    overlay_points = None

    if all_counts_rows:
        os.makedirs('results', exist_ok=True)
        df_counts = pd.DataFrame(all_counts_rows)

        # ===== Summary-Datei =====
        xlsx = os.path.join('results', f"doe_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if _HAS_XLSX:
            with pd.ExcelWriter(xlsx, engine='openpyxl') as writer:
                # Raw summary and optional markers
                df_counts.to_excel(writer, sheet_name='DOE_Summary', index=False)
                if all_marker_frames:
                    pd.concat(all_marker_frames, ignore_index=True).to_excel(writer, sheet_name='All_Markers', index=False)

                # Add Grand Total row in DOE_Summary
                ws = writer.sheets['DOE_Summary']
                n = len(df_counts) + 2
                ws.cell(row=n, column=1).value = 'Grand Total'
                ws.cell(row=n, column=4).value = f'=SUM(D2:D{n - 1})'
                ws.cell(row=n, column=5).value = f'=SUM(E2:E{n - 1})'
                ws.cell(row=n, column=6).value = f'=SUM(F2:F{n - 1})'

                # Nicely formatted pivot + heatmaps (Excel conditional formatting)
                try:
                    from openpyxl.formatting.rule import ColorScaleRule
                    from openpyxl.utils import get_column_letter

                    # 1) Overall pivot (Total across analyses)
                    agg_total = (df_counts.groupby(['Candle_Percent', 'MACD_Percent'], as_index=False)['Total'].sum())
                    if not agg_total.empty:
                        pv_all = (agg_total.pivot(index='Candle_Percent', columns='MACD_Percent', values='Total')
                                           .sort_index().sort_index(axis=1).fillna(0))
                        pv_all.to_excel(writer, sheet_name='DOE_Pivot_Total')
                        ws_pv = writer.sheets['DOE_Pivot_Total']
                        # Apply 3-color scale to numeric block
                        rows_n = pv_all.shape[0]
                        cols_n = pv_all.shape[1]
                        if rows_n > 0 and cols_n > 0:
                            r1, c1 = 2, 2
                            r2, c2 = 1 + rows_n, 1 + cols_n
                            ref = f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
                            ws_pv.conditional_formatting.add(ref, ColorScaleRule(start_type='min', start_color='93c47d',
                                                                                mid_type='percentile', mid_value=50, mid_color='ffd966',
                                                                                end_type='max', end_color='e06666'))
                            ws_pv.freeze_panes = ws_pv['B2']

                    # 2) Per-analysis heatmaps (Total, Classic, Hidden)
                    analyses = ["CBullDivg", "CBullDivg_x2", "HBullDivg", "HBearDivg"]
                    for a in analyses:
                        grp = df_counts[df_counts['Analysis'] == a]
                        if grp.empty:
                            continue
                        # Build three pivots
                        pivots = {}
                        for key in ['Total', 'Classic', 'Hidden']:
                            pv = (grp.pivot_table(index='Candle_Percent', columns='MACD_Percent', values=key, aggfunc='sum')
                                     .sort_index().sort_index(axis=1).fillna(0))
                            pivots[key] = pv

                        sheet_name = f"DOE_Heatmap_{a}"
                        # Write sequentially: header + matrix, with blank row between blocks
                        start_row = 1
                        with pd.option_context('display.float_format', lambda v: f"{v:.0f}"):
                            for title, pv in pivots.items():
                                if pv.empty:
                                    continue
                                # Write title
                                df_title = pd.DataFrame({f"{a} – {title}": []})
                                df_title.to_excel(writer, sheet_name=sheet_name, startrow=start_row - 1, index=False)
                                # Write pivot right below title
                                pv.to_excel(writer, sheet_name=sheet_name, startrow=start_row, startcol=0)
                                ws_hm = writer.sheets[sheet_name]
                                rows_n = pv.shape[0]
                                cols_n = pv.shape[1]
                                if rows_n > 0 and cols_n > 0:
                                    r1, c1 = start_row + 1, 2  # skip header row and index col
                                    r2, c2 = start_row + rows_n, 1 + cols_n
                                    ref = f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
                                    # Color palette per metric for visual separation
                                    if title == 'Total':
                                        start_col, mid_col, end_col = '93c47d', 'ffd966', 'e06666'
                                    elif title == 'Classic':
                                        start_col, mid_col, end_col = '9fc5e8', 'b4a7d6', '674ea7'
                                    else:  # Hidden
                                        start_col, mid_col, end_col = 'f9cb9c', 'f6b26b', 'cc0000'
                                    ws_hm.conditional_formatting.add(ref, ColorScaleRule(
                                        start_type='min', start_color=start_col,
                                        mid_type='percentile', mid_value=50, mid_color=mid_col,
                                        end_type='max', end_color=end_col))
                                    ws_hm.freeze_panes = ws_hm['B2']

                                # Next block two rows below
                                start_row += rows_n + 3
                # ---- Overview & Deltas sheets ----
                    from openpyxl.chart import BarChart, Reference
                    from openpyxl.utils import get_column_letter as _gcl

                    # Overview sheet aggregates
                    ov_sheet = 'Overview'
                    start_row = 0

                    # A) Totals per Analysis (table + bar chart)
                    totals_per_analysis = (df_counts.groupby('Analysis', as_index=False)['Total']
                                            .sum().sort_values('Total', ascending=False))
                    totals_per_analysis.to_excel(writer, sheet_name=ov_sheet, startrow=start_row, index=False)
                    ws_ov = writer.sheets[ov_sheet]
                    rows_n = len(totals_per_analysis) + 1
                    if rows_n > 1:
                        data_ref = Reference(ws_ov, min_col=2, min_row=start_row + 1, max_col=2, max_row=start_row + rows_n)
                        cats_ref = Reference(ws_ov, min_col=1, min_row=start_row + 2, max_row=start_row + rows_n)
                        ch = BarChart()
                        ch.title = 'Totals per Analysis'
                        ch.y_axis.title = 'Total markers'
                        ch.x_axis.title = 'Analysis'
                        ch.add_data(data_ref, titles_from_data=True)
                        ch.set_categories(cats_ref)
                        ch.height = 12
                        ch.width = 20
                        ws_ov.add_chart(ch, f"{_gcl(5)}{start_row + 2}")
                    start_row += rows_n + 2

                    # B) Analysis × MACD (table) and stacked chart by MACD
                    pivot_axm = (df_counts.pivot_table(index='MACD_Percent', columns='Analysis', values='Total', aggfunc='sum')
                                         .sort_index().fillna(0))
                    if not pivot_axm.empty:
                        pivot_axm.to_excel(writer, sheet_name=ov_sheet, startrow=start_row, startcol=0)
                        rows_n = pivot_axm.shape[0] + 1
                        cols_n = pivot_axm.shape[1] + 1
                        data_ref = Reference(ws_ov, min_col=2, min_row=start_row + 1,
                                             max_col=1 + pivot_axm.shape[1], max_row=start_row + rows_n)
                        cats_ref = Reference(ws_ov, min_col=1, min_row=start_row + 2, max_row=start_row + rows_n)
                        ch = BarChart()
                        ch.type = 'col'
                        ch.grouping = 'stacked'
                        ch.title = 'Totals by MACD (stacked by Analysis)'
                        ch.y_axis.title = 'Total markers'
                        ch.x_axis.title = 'MACD tolerance (%)'
                        ch.add_data(data_ref, titles_from_data=True)
                        ch.set_categories(cats_ref)
                        ch.height = 14
                        ch.width = 28
                        ws_ov.add_chart(ch, f"{_gcl(cols_n + 2)}{start_row + 2}")
                        start_row += rows_n + 2

                    # C) Analysis × Candle (table) and stacked chart by Candle
                    pivot_axc = (df_counts.pivot_table(index='Candle_Percent', columns='Analysis', values='Total', aggfunc='sum')
                                         .sort_index().fillna(0))
                    if not pivot_axc.empty:
                        pivot_axc.to_excel(writer, sheet_name=ov_sheet, startrow=start_row, startcol=0)
                        rows_n = pivot_axc.shape[0] + 1
                        cols_n = pivot_axc.shape[1] + 1
                        data_ref = Reference(ws_ov, min_col=2, min_row=start_row + 1,
                                             max_col=1 + pivot_axc.shape[1], max_row=start_row + rows_n)
                        cats_ref = Reference(ws_ov, min_col=1, min_row=start_row + 2, max_row=start_row + rows_n)
                        ch = BarChart()
                        ch.type = 'col'
                        ch.grouping = 'stacked'
                        ch.title = 'Totals by Candle (stacked by Analysis)'
                        ch.y_axis.title = 'Total markers'
                        ch.x_axis.title = 'Candle tolerance (%)'
                        ch.add_data(data_ref, titles_from_data=True)
                        ch.set_categories(cats_ref)
                        ch.height = 14
                        ch.width = 28
                        ws_ov.add_chart(ch, f"{_gcl(cols_n + 2)}{start_row + 2}")
                        start_row += rows_n + 2

                    # D) Simple matrices for quick scan: Analysis × MACD, Analysis × Candle
                    # Write Analysis as rows
                    mat_axm = (df_counts.pivot_table(index='Analysis', columns='MACD_Percent', values='Total', aggfunc='sum')
                                       .sort_index().sort_index(axis=1).fillna(0))
                    if not mat_axm.empty:
                        mat_axm.to_excel(writer, sheet_name=ov_sheet, startrow=start_row, startcol=0)
                        # conditional formatting
                        ws = writer.sheets[ov_sheet]
                        r1, c1 = start_row + 1, 2
                        r2, c2 = start_row + mat_axm.shape[0], 1 + mat_axm.shape[1]
                        ref = f"{_gcl(c1)}{r1}:{_gcl(c2)}{r2}"
                        ws.conditional_formatting.add(ref, ColorScaleRule(start_type='min', start_color='93c47d',
                                                                          mid_type='percentile', mid_value=50, mid_color='ffd966',
                                                                          end_type='max', end_color='e06666'))
                        start_row += mat_axm.shape[0] + 3

                    mat_axc = (df_counts.pivot_table(index='Analysis', columns='Candle_Percent', values='Total', aggfunc='sum')
                                       .sort_index().sort_index(axis=1).fillna(0))
                    if not mat_axc.empty:
                        mat_axc.to_excel(writer, sheet_name=ov_sheet, startrow=start_row, startcol=0)
                        ws = writer.sheets[ov_sheet]
                        r1, c1 = start_row + 1, 2
                        r2, c2 = start_row + mat_axc.shape[0], 1 + mat_axc.shape[1]
                        ref = f"{_gcl(c1)}{r1}:{_gcl(c2)}{r2}"
                        ws.conditional_formatting.add(ref, ColorScaleRule(start_type='min', start_color='93c47d',
                                                                          mid_type='percentile', mid_value=50, mid_color='ffd966',
                                                                          end_type='max', end_color='e06666'))

                    # E) Top-N parameter pairs (overall and per-analysis)
                    try:
                        top_n = int(os.getenv('DOE_TOP_N', 10))
                    except Exception:
                        top_n = 10

                    # Overall Top-N by Total
                    overall_pairs = (df_counts.groupby(['Candle_Percent', 'MACD_Percent'], as_index=False)
                                              .agg(Classic=('Classic','sum'), Hidden=('Hidden','sum'), Total=('Total','sum'))
                                              .sort_values('Total', ascending=False).head(top_n))
                    if not overall_pairs.empty:
                        # Add rank column first
                        overall_pairs = overall_pairs.reset_index(drop=True)
                        overall_pairs.insert(0, 'Rank', overall_pairs.index + 1)
                        title_df = pd.DataFrame({f"Top {top_n} parameter pairs – Overall": []})
                        title_df.to_excel(writer, sheet_name=ov_sheet, startrow=start_row, index=False)
                        overall_pairs.to_excel(writer, sheet_name=ov_sheet, startrow=start_row + 1, index=False)
                        start_row += len(overall_pairs) + 3

                    # Per-analysis Top-N (stacked vertically)
                    analyses = ["CBullDivg", "CBullDivg_x2", "HBullDivg", "HBearDivg"]
                    for a in analyses:
                        sub = df_counts[df_counts['Analysis'] == a]
                        if sub.empty:
                            continue
                        top_pairs = (sub.groupby(['Candle_Percent', 'MACD_Percent'], as_index=False)
                                         .agg(Classic=('Classic','sum'), Hidden=('Hidden','sum'), Total=('Total','sum'))
                                         .sort_values('Total', ascending=False).head(top_n))
                        if top_pairs.empty:
                            continue
                        top_pairs = top_pairs.reset_index(drop=True)
                        top_pairs.insert(0, 'Rank', top_pairs.index + 1)
                        title_df = pd.DataFrame({f"Top {top_n} parameter pairs – {a}": []})
                        title_df.to_excel(writer, sheet_name=ov_sheet, startrow=start_row, index=False)
                        top_pairs.to_excel(writer, sheet_name=ov_sheet, startrow=start_row + 1, index=False)
                        start_row += len(top_pairs) + 3

                    # Deltas vs. baseline (first DOE row as baseline)
                    try:
                        b_candle = float(prm.iloc[0]['candle_percent'])
                        b_macd = float(prm.iloc[0]['macd_percent'])
                        del_sheet = 'DOE_Deltas'
                        start_row = 1
                        analyses = ["CBullDivg", "CBullDivg_x2", "HBullDivg", "HBearDivg"]
                        for a in analyses:
                            base_total_row = df_counts[(df_counts['Candle_Percent'] == b_candle) &
                                                       (df_counts['MACD_Percent'] == b_macd) &
                                                       (df_counts['Analysis'] == a)]
                            if base_total_row.empty:
                                continue
                            base_total = float(base_total_row['Total'].iloc[0])
                            grp = (df_counts[df_counts['Analysis'] == a]
                                            .groupby(['Candle_Percent', 'MACD_Percent'], as_index=False)
                                            .sum(numeric_only=True))
                            pv_total = (grp.pivot(index='Candle_Percent', columns='MACD_Percent', values='Total')
                                          .sort_index().sort_index(axis=1).fillna(0))
                            pv_delta = pv_total - base_total
                            # Write title and delta matrix
                            df_title = pd.DataFrame({f"Delta to baseline ({b_candle}, {b_macd}) – {a}": []})
                            df_title.to_excel(writer, sheet_name=del_sheet, startrow=start_row - 1, index=False)
                            pv_delta.to_excel(writer, sheet_name=del_sheet, startrow=start_row, startcol=0)
                            ws_del = writer.sheets[del_sheet]
                            rows_n = pv_delta.shape[0]
                            cols_n = pv_delta.shape[1]
                            if rows_n > 0 and cols_n > 0:
                                r1, c1 = start_row + 1, 2
                                r2, c2 = start_row + rows_n, 1 + cols_n
                                ref = f"{_gcl(c1)}{r1}:{_gcl(c2)}{r2}"
                                ws_del.conditional_formatting.add(ref, ColorScaleRule(
                                    start_type='min', start_color='9fc5e8',
                                    mid_type='percentile', mid_value=50, mid_color='ffffff',
                                    end_type='max', end_color='e06666'))
                                ws_del.freeze_panes = ws_del['B2']
                            start_row += rows_n + 3
                    except Exception:
                        pass

                except Exception as e:
                    # Formatting is best-effort; keep file creation robust
                    print(f"Warning: could not add Excel pivots/heatmaps: {e}")

            print(f"DOE summary exported to {xlsx}")
        else:
            csv = xlsx.replace('.xlsx', '.csv')
            df_counts.to_csv(csv, index=False)
            print(f"DOE summary exported to {csv}")

        # ===== Optional robust scoring via walk-forward CV =====
        try:
            n_splits = int(os.getenv('DOE_WF_SPLITS', '0'))
        except Exception:
            n_splits = 0
        if n_splits and n_splits > 1 and 'date' in df.columns:
            print(f"Walk-forward CV with {n_splits} folds for robust scoring...")
            dates = pd.to_datetime(df['date'])
            cut_idx = np.linspace(0, len(df), n_splits + 1, dtype=int)
            lam = float(os.getenv('DOE_ROBUST_LAMBDA', '1.0'))
            analyses = ["CBullDivg", "CBullDivg_x2", "HBullDivg", "HBearDivg"]
            per_type_points = {a: [] for a in analyses}

            for a in analyses:
                atype_char = {'CBullDivg':'a','CBullDivg_x2':'b','HBearDivg':'c','HBullDivg':'d'}[a]
                for _, ctol, mtol in indexed_params:
                    totals = []
                    for k in range(n_splits):
                        lo, hi = cut_idx[k], cut_idx[k+1]
                        df_slice = df.iloc[lo:hi].copy()
                        res = run_analysis(df_slice, atype_char, window, ctol, mtol)
                        counts_k = export_markers_to_csv(df_slice, f"_tmp_cv_{a}_{k}.csv", res, ctol, mtol)
                        totals.append(counts_k.get(a, {}).get('total', 0))
                    mean_v = float(np.mean(totals)) if totals else 0.0
                    std_v = float(np.std(totals, ddof=1)) if len(totals) > 1 else 0.0
                    score = mean_v - lam * std_v
                    per_type_points[a].append({
                        'Candle_Percent': ctol,
                        'MACD_Percent': mtol,
                        'mean': mean_v,
                        'std': std_v,
                        'score': score,
                    })

            # Top-N and Pareto front overlay
            def _pareto_front(pts):
                front = []
                for i, p in enumerate(pts):
                    dom = False
                    for j, q in enumerate(pts):
                        if j == i:
                            continue
                        if (q['mean'] >= p['mean']) and (q['std'] <= p['std']) and ((q['mean'] > p['mean']) or (q['std'] < p['std'])):
                            dom = True
                            break
                    if not dom:
                        front.append(p)
                return front

            try:
                topn = int(os.getenv('DOE_TOPN_OVERLAY', '5'))
            except Exception:
                topn = 5
            overlay_points = {}
            for a, pts in per_type_points.items():
                if not pts:
                    continue
                top = sorted(pts, key=lambda p: p['score'], reverse=True)[:topn]
                for t in top:
                    t['label'] = f"S:{t['score']:.1f}"
                pareto = _pareto_front(pts)
                for p in pareto:
                    p['label'] = (p.get('label') or '') + ' P'
                overlay_points[a] = top + pareto

        # ===== Interaktive DOE-Plots =====
        if make_total:
            _doe_plot_total(df_counts, asset_label=asset_label)
        _doe_plot_by_type(df_counts, single_pages=make_single, asset_label=asset_label, overlay_points=overlay_points)


# -------------------------------------------------
# MAIN
# -------------------------------------------------
if __name__ == "__main__":
    path = get_input_file()
    if not path:
        print("No file selected. Exiting.")
        raise SystemExit(0)

    atype = get_analysis_type()

    # Parameter nur, wenn nicht DOE
    if atype not in ('f', 'g', 'h'):
        candle_percent, macd_percent, variant2 = get_analysis_parameters()

    print(f"Loading data from: {path}")
    if path.lower().endswith('.parquet'):
        df = pd.read_parquet(path)
    else:
        df = pd.read_csv(path, low_memory=False)

    print(f"Loaded {len(df)} rows")

    # Spalte 'date' sicherstellen
    if 'date' not in df.columns and 'timestamp' in df.columns:
        df['date'] = df['timestamp']

    df['date'] = pd.to_datetime(df['date'], utc=True, errors='coerce')
    df = df.dropna(subset=['date']).sort_values('date').reset_index(drop=True)
    # Echo full available span after load
    try:
        _min_dt = df['date'].min()
        _max_dt = df['date'].max()
        print(f"Asset timespan: {_min_dt} → {_max_dt}")
    except Exception:
        pass

    # Asset label for plot titles
    def _asset_label_from(path_in, df_in):
        base = os.path.splitext(os.path.basename(path_in))[0]
        try:
            if 'symbol' in df_in.columns and df_in['symbol'].dropna().nunique() == 1:
                sym = str(df_in['symbol'].dropna().iloc[0])
                return f"{sym} ({base})"
        except Exception:
            pass
        return base

    asset_label = _asset_label_from(path, df)
    print(f"Asset: {asset_label}")

    print("Initializing indicators...")
    Initialize_RSI_EMA_MACD(df)
    Local_Max_Min(df)

    window = 5

    if atype not in ('f', 'g', 'h'):
        print(f"Running analysis with parameters: window={window}, candle_tol={candle_percent}, macd_tol={macd_percent}")
        analysis_results = run_analysis(df, atype, window, candle_percent, macd_percent)
        out_csv = f"markers_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        counts = export_markers_to_csv(df, out_csv, analysis_results, candle_percent, macd_percent)

        df2 = None
        res2 = None
        counts2 = None
        if variant2:
            print(f"\nRunning second variant: candle={variant2[0]}, macd={variant2[1]}")
            df2 = df.copy()
            res2 = run_analysis(df2, atype, window, variant2[0], variant2[1])
            out_csv2 = f"variant2_{variant2[0]}_{variant2[1]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            counts2 = export_markers_to_csv(df2, out_csv2, res2, variant2[0], variant2[1])

        # Plot (Index auf 'date')
        df = df.set_index('date')
        if df2 is not None:
            df2 = df2.set_index('date')
        plot_with_plotly(df, analysis_results, counts, df2, res2, counts2, asset_label=asset_label)

    elif atype == 'f':
        # Deine Vorgabe: NUR 2×2-HTML mit allen vier Divergenztypen
        run_doe_analysis(df, make_total=False, make_single=False, asset_label=asset_label)
    else:
        # Backtesting path
        # Collect markers CSVs from results folder
        try:
            candidates = []
            os.makedirs('results', exist_ok=True)
            for fn in os.listdir('results'):
                if fn.lower().endswith('.csv') and 'markers' in fn.lower():
                    candidates.append(os.path.join('results', fn))
            candidates.sort()
            selected = []
            if atype == 'h':
                selected = [p for p in candidates if os.path.basename(p).startswith('doe_markers_')]
                if not selected:
                    print("No doe_markers_*.csv found in results/. Run DOE first.")
                    raise SystemExit(0)
            else:
                print("\nAvailable markers CSV files in results/:")
                for i, p in enumerate(candidates, 1):
                    print(f"{i}: {p}")
                print("0: Use ALL listed")
                sel = input("Select file number (or 0 for ALL): ").strip()
                if sel.isdigit() and int(sel) > 0 and int(sel) <= len(candidates):
                    selected = [candidates[int(sel) - 1]]
                else:
                    selected = candidates
            if not selected:
                print("No markers CSV found in results/. Please run an analysis first.")
                raise SystemExit(0)

            # Optional: restrict trading timespan
            s_dt, e_dt = select_backtest_timespan(df)
            if s_dt and e_dt:
                print(f"Selected backtest span: {s_dt} to {e_dt}")
                before = len(df)
                df = df[(df['date'] >= s_dt) & (df['date'] <= e_dt)].reset_index(drop=True)
                print(f"Rows after span filter: {len(df)} (from {before})")
                # Sanity echo of new actual span
                try:
                    print(f"Applied backtest span: {df['date'].min()} to {df['date'].max()}")
                except Exception:
                    pass
            else:
                # Use full span
                s_dt = pd.to_datetime(df['date']).min()
                e_dt = pd.to_datetime(df['date']).max()

            # Load markers
            mks = []
            for p in selected:
                try:
                    dfm = pd.read_csv(p)
                    if 'Date' in dfm.columns:
                        dfm['SourceCSV'] = os.path.basename(p)
                        if s_dt and e_dt:
                            # prefilter markers into selected span
                            tmp = dfm.copy()
                            tmp['Date'] = pd.to_datetime(tmp['Date'], utc=True, errors='coerce')
                            tmp = tmp[(tmp['Date'] >= s_dt) & (tmp['Date'] <= e_dt)]
                            mks.append(tmp)
                        else:
                            mks.append(dfm)
                except Exception as e:
                    print(f"Warning: cannot load {p}: {e}")
            if not mks:
                print("No valid markers loaded. Exiting.")
                raise SystemExit(0)
            markers_df = pd.concat(mks, ignore_index=True)

            # Prompt for backtest parameters and run
            bt_params = backtest_prompt_params()
            # Attach selected backtest timespan to params for reporting
            try:
                bt_params.backtest_start = s_dt
                bt_params.backtest_end = e_dt
            except Exception:
                pass
            # Robust printing (avoid issues if attributes missing)
            ps = bt_params
            parts = [
                f"risk={getattr(ps, 'risk_pct', 'n/a')}%",
                f"stop={getattr(ps, 'stop_pct', 'n/a')}%",
                f"tp={getattr(ps, 'tp_pct', 'n/a')}%",
                f"fee/side={getattr(ps, 'fee_pct', 'n/a')}%",
                f"slippage={getattr(ps, 'slippage_pct', 'n/a')}%",
                f"max_pos%={getattr(ps, 'max_position_value_pct', 'n/a')}%",
                f"eq_cap={getattr(ps, 'equity_cap', 'n/a')}",
                f"time_stop={getattr(ps, 'time_stop_bars', 'n/a')}",
                f"single_pos={getattr(ps, 'single_position_mode', 'n/a')}",
                f"span=[{s_dt} → {e_dt}]",
            ]
            print("\nBacktesting with: " + ", ".join(parts))
            results = run_backtest(df.copy(), markers_df, bt_params)

            # Console summary
            summ = results.get('summary', pd.DataFrame())
            if not summ.empty:
                s = summ.iloc[0].to_dict()
                print("\n=== Backtest Summary ===")
                for k, v in s.items():
                    print(f"{k}: {v}")
            else:
                print("No trades generated for the selected markers/params.")

            # XLSX export
            base_name = 'backtest_doe_combined' if atype == 'h' else 'backtest'
            # Append span to filename: sYYYYMMDD[_HHMM]-eYYYYMMDD[_HHMM]
            def _fmt_span(ts):
                try:
                    # include time if intraday
                    return pd.to_datetime(ts).strftime('%Y%m%d_%H%M') if (getattr(pd.to_datetime(ts), 'hour', 0) != 0 or getattr(pd.to_datetime(ts), 'minute', 0) != 0) else pd.to_datetime(ts).strftime('%Y%m%d')
                except Exception:
                    return 'span'
            span_tag = f"s{_fmt_span(s_dt)}-e{_fmt_span(e_dt)}"
            out_xlsx = os.path.join('results', f"{base_name}_{span_tag}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            export_backtest_xlsx(results, out_xlsx)
            print(f"Backtest report saved to: {out_xlsx}")
        except (EOFError, KeyboardInterrupt):
            print("Backtest cancelled by user.")
