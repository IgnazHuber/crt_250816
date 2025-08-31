import os
from datetime import datetime
from typing import List, Dict, Any, Tuple

import numpy as np
import pandas as pd

try:
    import openpyxl  # noqa: F401
    _HAS_XLSX = True
except Exception:
    _HAS_XLSX = False


def _parse_grid(env_name: str, default: str) -> List[float]:
    raw = os.getenv(env_name, default)
    out = []
    for part in str(raw).replace(';', ',').split(','):
        part = part.strip()
        if not part:
            continue
        try:
            out.append(float(part))
        except Exception:
            pass
    return out


def run_secondary_sweep(df: pd.DataFrame,
                        asset_label: str,
                        asset_tag: str,
                        start_dt: pd.Timestamp = None,
                        end_dt: pd.Timestamp = None,
                        doe_params_file: str = "doe_parameters_example.csv") -> str:
    """
    Stage-B optimizer: For shortlisted indicator cells (from DOE params file),
    sweep trading params (risk, stop, tp) with walk-forward CV and export an XLSX.

    Env controls (with defaults):
      SS_WF_SPLITS=3
      SS_GRID_RISK="0.5,1,2,3,4,5"
      SS_GRID_STOP="2,4,6,8,10,12"
      SS_GRID_TP="5,10,15,20,25,30"
      SS_LAMBDA=1.0 (robust score lambda)
      SS_MIN_TRADES=20
      SS_MAX_DD=50 (percent)
      SS_MIN_PF=1.2
      SS_TOPN=15
      SS_SCORE_METHOD=robust|scalar (robust default)
    """
    # Lazy imports to avoid circular import at module import-time
    from Backtest_Divergences import BacktestParams, backtest as run_backtest
    from Mainframe_RTv250829_mplfinance import run_analysis, generate_markers_df

    # Parse grids and settings
    try:
        n_splits = int(os.getenv('SS_WF_SPLITS', '3'))
    except Exception:
        n_splits = 3
    risk_grid = _parse_grid('SS_GRID_RISK', '0.5,1,2,3,4,5')
    stop_grid = _parse_grid('SS_GRID_STOP', '2,4,6,8,10,12')
    tp_grid = _parse_grid('SS_GRID_TP', '5,10,15,20,25,30')
    try:
        lam = float(os.getenv('SS_LAMBDA', '1.0'))
    except Exception:
        lam = 1.0
    try:
        min_trades = int(os.getenv('SS_MIN_TRADES', '20'))
    except Exception:
        min_trades = 20
    try:
        max_dd_cap = float(os.getenv('SS_MAX_DD', '50'))
    except Exception:
        max_dd_cap = 50.0
    try:
        min_pf = float(os.getenv('SS_MIN_PF', '1.2'))
    except Exception:
        min_pf = 1.2
    score_method = os.getenv('SS_SCORE_METHOD', 'robust').lower()
    try:
        topn = int(os.getenv('SS_TOPN', '15'))
    except Exception:
        topn = 15

    # Indicator candidates from DOE params CSV
    try:
        prm = pd.read_csv(doe_params_file)
        if not {'candle_percent', 'macd_percent'}.issubset(prm.columns):
            raise ValueError('DOE params file must have candle_percent, macd_percent')
        indicator_cells = [(float(r['candle_percent']), float(r['macd_percent'])) for _, r in prm.iterrows()]
    except Exception as e:
        # fallback to a small default grid
        indicator_cells = [(0.1, 3.25), (0.2, 2.0), (0.5, 5.0)]

    # Walk-forward splits
    if n_splits < 2:
        cut_idx = np.array([0, len(df)], dtype=int)
    else:
        cut_idx = np.linspace(0, len(df), n_splits + 1, dtype=int)

    rows = []
    # Fee/slippage defaults from env
    try:
        fee_def = float(os.getenv('BT_FEE_PCT', '0.0'))
    except Exception:
        fee_def = 0.0
    try:
        slip_def = float(os.getenv('BT_SLIPPAGE_PCT', '0.0'))
    except Exception:
        slip_def = 0.0

    window = 5
    print(f"Secondary sweep: {len(indicator_cells)} indicator cells × {len(risk_grid)*len(stop_grid)*len(tp_grid)} trading tuples across {n_splits} folds…")

    # Parallelize over parameter tuples
    from concurrent.futures import ThreadPoolExecutor, as_completed
    param_tuples = [(c, m, r, s, t) for (c, m) in indicator_cells for r in risk_grid for s in stop_grid for t in tp_grid]
    total = len(param_tuples)
    done = 0
    last_pct = -1

    def _worker(candle_tol, macd_tol, risk_pct, stop_pct, tp_pct):
        pnl_vals = []
        pf_vals = []
        wr_vals = []
        dd_vals = []
        tr_vals = []
        start_vals = []
        for k in range(max(1, n_splits)):
            lo = cut_idx[k] if n_splits >= 2 else 0
            hi = cut_idx[k+1] if n_splits >= 2 else len(df)
            df_slice = df.iloc[lo:hi].copy()
            res = run_analysis(df_slice, 'e', window, candle_tol, macd_tol)
            mk_df, _ = generate_markers_df(df_slice, res, candle_tol, macd_tol, asset_tag=asset_tag)
            if mk_df is None or mk_df.empty:
                # No signals in this fold; skip
                continue
            params = BacktestParams(
                risk_pct=risk_pct,
                stop_pct=stop_pct,
                tp_pct=tp_pct,
                fee_pct=fee_def,
                slippage_pct=slip_def,
                conservative_sizing=True,
            )
            bt_res = run_backtest(df_slice.copy(), mk_df.copy(), params)
            summ = bt_res.get('summary')
            if summ is None or summ.empty:
                continue
            s = summ.iloc[0]
            start_vals.append(float(s.get('Starting_Capital', 10000.0)))
            pnl_vals.append(float(s.get('Total_PnL', 0.0)))
            pf_vals.append(float(s.get('Profit_Factor', 0.0)) if pd.notna(s.get('Profit_Factor', np.nan)) else 0.0)
            wr_vals.append(float(s.get('Win_Rate_%', 0.0)))
            dd_vals.append(float(s.get('Max_Drawdown_%', 0.0)) if pd.notna(s.get('Max_Drawdown_%', np.nan)) else 0.0)
            tr_vals.append(float(s.get('Trades', 0.0)))

        if not pnl_vals:
            return None
        mean_pnl = float(np.mean(pnl_vals))
        std_pnl = float(np.std(pnl_vals, ddof=1)) if len(pnl_vals) > 1 else 0.0
        mean_pf = float(np.mean(pf_vals)) if pf_vals else 0.0
        mean_wr = float(np.mean(wr_vals)) if wr_vals else 0.0
        mean_dd = float(np.mean(dd_vals)) if dd_vals else 0.0
        mean_tr = float(np.mean(tr_vals)) if tr_vals else 0.0
        mean_start = float(np.mean(start_vals)) if start_vals else 10000.0
        mean_final_equity = mean_start + mean_pnl

        feasible = (mean_tr >= min_trades) and (mean_pf >= min_pf) and (mean_dd <= max_dd_cap)
        if score_method == 'robust':
            score = (mean_pnl - lam * std_pnl) if feasible else -1e18
        else:
            score = (mean_pnl) + 0.5 * (mean_pf) - 0.5 * (mean_dd)
            if not feasible:
                score -= 1e6

        return {
            'Candle_Percent': candle_tol,
            'MACD_Percent': macd_tol,
            'Risk_%': risk_pct,
            'Stop_%': stop_pct,
            'TP_%': tp_pct,
            'Mean_PnL': mean_pnl,
            'Std_PnL': std_pnl,
            'Mean_PF': mean_pf,
            'Mean_WR_%': mean_wr,
            'Mean_MaxDD_%': mean_dd,
            'Mean_Trades': mean_tr,
            'Score': score,
            'Feasible': feasible,
            'Mean_Final_Equity': mean_final_equity,
        }

    try:
        max_workers = int(os.getenv('SS_MAX_WORKERS', '0')) or os.cpu_count() or 4
    except Exception:
        max_workers = os.cpu_count() or 4

    print(f"Launching {total} parameter tasks with {max_workers} workers…")
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = [ex.submit(_worker, *pt) for pt in param_tuples]
        for f in as_completed(futs):
            done += 1
            r = f.result()
            if r is not None:
                rows.append(r)
            pct = int(done * 100 / total)
            if pct != last_pct and pct % 2 == 0:
                print(f"\rProgress: {done}/{total} ({pct}%)", end="", flush=True)
                last_pct = pct
    print()

    if not rows:
        # Export info-only workbook explaining no results
        out_name = f"{asset_tag}_secondary_sweep_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        out_path = os.path.join('results', out_name)
        os.makedirs('results', exist_ok=True)
        if _HAS_XLSX:
            with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                pd.DataFrame([{
                    'Asset_Tag': asset_tag,
                    'Asset_Label': asset_label,
                    'Start': str(start_dt) if start_dt is not None else '',
                    'End': str(end_dt) if end_dt is not None else '',
                    'Note': 'No signals across folds for the tested parameter grid.'
                }]).to_excel(writer, sheet_name='Info', index=False)
        print("Secondary sweep produced no results.")
        return out_path

    df_res = pd.DataFrame(rows)
    df_res.sort_values('Score', ascending=False, inplace=True)

    # Pareto front on (Mean_PnL, Mean_MaxDD_%)
    def pareto(points: List[Dict[str, Any]]):
        front = []
        for i, p in enumerate(points):
            dom = False
            for j, q in enumerate(points):
                if i == j: continue
                if (q['Mean_PnL'] >= p['Mean_PnL']) and (q['Mean_MaxDD_%'] <= p['Mean_MaxDD_%']) and ((q['Mean_PnL'] > p['Mean_PnL']) or (q['Mean_MaxDD_%'] < p['Mean_MaxDD_%'])):
                    dom = True; break
            if not dom:
                front.append(p)
        return front

    pareto_points = pareto(df_res.to_dict(orient='records'))

    # Export XLSX
    out_name = f"{asset_tag}_secondary_sweep_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = os.path.join('results', out_name)
    os.makedirs('results', exist_ok=True)

    if not _HAS_XLSX:
        df_res.to_csv(out_path.replace('.xlsx', '.csv'), index=False)
        print(f"Secondary sweep results saved to {out_path.replace('.xlsx', '.csv')}")
        return out_path

    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.utils import get_column_letter as _gcl

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        # Overview top-N
        top = df_res.head(topn).copy()
        top.to_excel(writer, sheet_name='Overview', index=False)

        # All results (optional large)
        df_res.to_excel(writer, sheet_name='All_Results', index=False)

        # Pareto table
        df_par = pd.DataFrame(pareto_points)
        if not df_par.empty:
            df_par.to_excel(writer, sheet_name='Pareto', index=False)
            ws = writer.sheets['Pareto']
            r = ws.max_row
            c = ws.max_column
            # Build scatter chart: X=Mean_MaxDD_%, Y=Mean_PnL
            # Assumes columns: Mean_PnL and Mean_MaxDD_% exist
            headers = list(df_par.columns)
            try:
                x_col = headers.index('Mean_MaxDD_%') + 1
                y_col = headers.index('Mean_PnL') + 1
                chart = ScatterChart()
                chart.title = 'Pareto: PnL vs MaxDD'
                chart.x_axis.title = 'Mean MaxDD % (lower better)'
                chart.y_axis.title = 'Mean PnL (higher better)'
                xref = Reference(ws, min_col=x_col, min_row=2, max_row=r)
                yref = Reference(ws, min_col=y_col, min_row=2, max_row=r)
                series = Series(yref, xref, title_from_data=False, title='Pareto')
                chart.series.append(series)
                ws.add_chart(chart, f"{_gcl(c+2)}2")
            except Exception:
                pass

        # Info sheet
        info = [{
            'Asset_Tag': asset_tag,
            'Asset_Label': asset_label,
            'Start': str(start_dt) if start_dt is not None else '',
            'End': str(end_dt) if end_dt is not None else '',
            'WF_Splits': n_splits,
            'Lambda': lam,
            'Min_Trades': min_trades,
            'MaxDD_Cap_%': max_dd_cap,
            'Min_PF': min_pf,
            'Score_Method': score_method,
        }]
        pd.DataFrame(info).to_excel(writer, sheet_name='Info', index=False)

    print(f"Secondary sweep report saved to: {out_path}")
    return out_path


def _asset_tag_from(path_in: str) -> str:
    base = os.path.splitext(os.path.basename(path_in))[0]
    parts = base.split('_')
    if len(parts) >= 2:
        return f"{parts[0]}_{parts[1]}"
    return parts[0]


def _load_asset_df(path: str) -> pd.DataFrame:
    import pandas as pd
    from Initialize_RSI_EMA_MACD import Initialize_RSI_EMA_MACD
    from Local_Maximas_Minimas import Local_Max_Min
    if path.lower().endswith('.parquet'):
        df = pd.read_parquet(path)
    else:
        df = pd.read_csv(path, low_memory=False)
    if 'date' not in df.columns and 'timestamp' in df.columns:
        df['date'] = df['timestamp']
    df['date'] = pd.to_datetime(df['date'], utc=True, errors='coerce')
    df = df.dropna(subset=['date']).sort_values('date').reset_index(drop=True)
    Initialize_RSI_EMA_MACD(df)
    Local_Max_Min(df)
    return df


def run_secondary_sweep_multi(paths: List[str],
                              span_start: str = None,
                              span_end: str = None,
                              doe_params_file: str = "doe_parameters_example.csv") -> str:
    """Run secondary sweep for multiple assets and build a combined report selecting the best robust row per asset.
    Optional span_start/span_end as ISO strings to apply to each asset; defaults to full span.
    """
    results = []
    reports = []
    for p in paths:
        try:
            df = _load_asset_df(p)
            s_dt = pd.to_datetime(span_start, utc=True) if span_start else df['date'].min()
            e_dt = pd.to_datetime(span_end, utc=True) if span_end else df['date'].max()
            df = df[(df['date'] >= s_dt) & (df['date'] <= e_dt)].reset_index(drop=True)
            tag = _asset_tag_from(p)
            label = os.path.splitext(os.path.basename(p))[0]
            report = run_secondary_sweep(df, label, tag, s_dt, e_dt, doe_params_file)
            reports.append(report)
            # Read back All_Results to pick best robust row
            if _HAS_XLSX and report and os.path.isfile(report):
                try:
                    df_all = pd.read_excel(report, sheet_name='All_Results')
                except Exception:
                    df_all = pd.read_csv(report.replace('.xlsx', '.csv')) if os.path.isfile(report.replace('.xlsx', '.csv')) else pd.DataFrame()
            else:
                df_all = pd.DataFrame()
            if not df_all.empty:
                best = df_all.sort_values('Score', ascending=False).head(1).copy()
                best.insert(0, 'Asset_Tag', tag)
                best.insert(1, 'Asset_Label', label)
                results.append(best)
        except Exception as e:
            print(f"Secondary sweep failed for {p}: {e}")
            continue

    if not results:
        print("No combined results produced.")
        return ''

    df_best = pd.concat(results, ignore_index=True)
    # Frequency column from tag (second token)
    def _freq_from_tag(tag: str) -> str:
        try:
            return str(tag).split('_')[1]
        except Exception:
            return ''
    df_best['Frequency'] = df_best['Asset_Tag'].map(_freq_from_tag)

    out_combined = os.path.join('results', f"combined_secondary_sweep_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    from openpyxl.chart import BarChart, Reference
    with pd.ExcelWriter(out_combined, engine='openpyxl') as writer:
        df_best.to_excel(writer, sheet_name='Best_By_Asset', index=False)
        # Top-N assets by Score chart
        try:
            topn = int(os.getenv('COMBINED_TOPN', '15'))
        except Exception:
            topn = 15
        df_top_assets = df_best.sort_values('Score', ascending=False).head(topn)
        if not df_top_assets.empty:
            df_top_assets.to_excel(writer, sheet_name='Top_Assets', index=False)
            ws = writer.sheets['Top_Assets']
            r = ws.max_row
            headers = list(df_top_assets.columns)
            # Prefer Asset_Label for x-axis, fallback to Asset_Tag
            try:
                label_col = headers.index('Asset_Label') + 1
            except ValueError:
                label_col = headers.index('Asset_Tag') + 1 if 'Asset_Tag' in headers else 1
            score_col = headers.index('Score') + 1 if 'Score' in headers else None
            if score_col:
                ch = BarChart()
                ch.title = f'Top {topn} Robust Scores by Asset'
                ch.y_axis.title = 'Score'
                ch.x_axis.title = 'Asset'
                data_ref = Reference(ws, min_col=score_col, min_row=1, max_row=r)
                cats_ref = Reference(ws, min_col=label_col, min_row=2, max_row=r)
                ch.add_data(data_ref, titles_from_data=True)
                ch.set_categories(cats_ref)
                ws.add_chart(ch, 'J2')
        # Best per frequency
        if 'Frequency' in df_best.columns:
            idx = df_best.groupby('Frequency')['Score'].idxmax()
            df_freq = df_best.loc[idx].sort_values('Frequency')
            df_freq.to_excel(writer, sheet_name='Best_By_Frequency', index=False)
            # Add simple bar chart of top scores per frequency
            ws = writer.sheets['Best_By_Frequency']
            r = ws.max_row
            # Find columns for Frequency and Score
            headers = list(df_freq.columns)
            try:
                freq_col = headers.index('Frequency') + 1
                score_col = headers.index('Score') + 1
                chart = BarChart()
                chart.title = 'Top Robust Scores by Frequency'
                chart.y_axis.title = 'Score'
                chart.x_axis.title = 'Frequency'
                data_ref = Reference(ws, min_col=score_col, min_row=1, max_row=r)
                cats_ref = Reference(ws, min_col=freq_col, min_row=2, max_row=r)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                ws.add_chart(chart, 'J2')
            except Exception:
                pass
        # Reports list
        pd.DataFrame({'Report': reports}).to_excel(writer, sheet_name='Reports', index=False)
    print(f"Combined secondary sweep saved to: {out_combined}")
    return out_combined
