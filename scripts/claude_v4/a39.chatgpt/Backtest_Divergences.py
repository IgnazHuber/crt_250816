import os
from dataclasses import dataclass
from datetime import datetime
from typing import List, Dict, Any

import numpy as np
import pandas as pd

try:
    import openpyxl  # noqa: F401
    _HAS_XLSX = True
except Exception:
    _HAS_XLSX = False


@dataclass
class BacktestParams:
    starting_capital: float = 10000.0
    risk_pct: float = 5.0   # percent of equity
    stop_pct: float = 5.0   # trailing stop percent
    tp_pct: float = 25.0    # take profit percent
    fee_pct: float = 0.0    # per side fee percent of notional
    slippage_pct: float = 0.0  # adverse slippage percent
    max_position_value_pct: float = 100.0  # cap position notional as % of equity
    equity_cap: float = 0.0  # if >0, cap effective equity for sizing
    time_stop_bars: int = 0  # if >0, exit after N bars
    single_position_mode: bool = False  # if True, do not open if another is open
    allow_shorts: bool = True
    backtest_start: object = None
    backtest_end: object = None
    conservative_sizing: bool = False  # if True, use entry_fill for stop distance
    max_open_positions: int = 0  # 0 = unlimited
    max_open_longs: int = 0      # 0 = unlimited
    max_open_shorts: int = 0     # 0 = unlimited


def prompt_params() -> BacktestParams:
    def _ask(prompt: str, default: float) -> float:
        try:
            s = input(f"{prompt} (default {default}%): ").strip()
            return float(s) if s else float(default)
        except (EOFError, KeyboardInterrupt):
            return float(default)
        except Exception:
            print("Invalid input. Using default.")
            return float(default)

    print("\nEnter backtest parameters (press Enter for defaults):")
    risk = _ask("Risk per trade %", 5.0)
    sl = _ask("Trailing stop %", 5.0)
    tp = _ask("Take profit %", 25.0)
    # Global defaults via env vars
    try:
        fee_def = float(os.getenv('BT_FEE_PCT', '0.0'))
    except Exception:
        fee_def = 0.0
    try:
        slip_def = float(os.getenv('BT_SLIPPAGE_PCT', '0.0'))
    except Exception:
        slip_def = 0.0
    fee = _ask(f"Fee per side % (0 for none)", fee_def)
    slip = _ask("Slippage % (adverse)", slip_def)

    # Advanced controls with env defaults
    try:
        max_pos_def = float(os.getenv('BT_MAX_POS_VAL_PCT', '5.0'))
    except Exception:
        max_pos_def = 5.0
    try:
        eq_cap_def = float(os.getenv('BT_EQUITY_CAP', '0.0'))
    except Exception:
        eq_cap_def = 0.0
    try:
        tstop_def = float(os.getenv('BT_TIME_STOP_BARS', '0'))
    except Exception:
        tstop_def = 0
    sp_def = os.getenv('BT_SINGLE_POSITION', 'false').lower() in ('1','true','yes','y')
    cons_def = os.getenv('BT_CONSERVATIVE_SIZING', 'false').lower() in ('1','true','yes','y')
    try:
        max_pos_open_def = int(os.getenv('BT_MAX_OPEN', '0'))
    except Exception:
        max_pos_open_def = 0
    try:
        max_open_longs_def = int(os.getenv('BT_MAX_OPEN_LONGS', '0'))
    except Exception:
        max_open_longs_def = 0
    try:
        max_open_shorts_def = int(os.getenv('BT_MAX_OPEN_SHORTS', '0'))
    except Exception:
        max_open_shorts_def = 0

    max_pos = _ask("Max position notional % of equity (default 5.0)", max_pos_def)
    eq_cap = _ask("Equity cap for sizing (0=none)", eq_cap_def)
    try:
        tstop = int(_ask("Time stop bars (0=off)", tstop_def))
    except Exception:
        tstop = 0
    try:
        sp_in = input(f"Single-position mode? (y/N) [default {'Y' if sp_def else 'N'}]: ").strip().lower()
        single_pos = sp_def if sp_in == '' else (sp_in in ('y','yes','true','1'))
    except (EOFError, KeyboardInterrupt):
        single_pos = sp_def

    try:
        cons_in = input(f"Conservative sizing (use entry fill for stop distance)? (y/N) [default {'Y' if cons_def else 'N'}]: ").strip().lower()
        conservative = cons_def if cons_in == '' else (cons_in in ('y','yes','true','1'))
    except (EOFError, KeyboardInterrupt):
        conservative = cons_def
    # caps for concurrent positions
    def _ask_int(prompt_text, default_val):
        try:
            s = input(f"{prompt_text} (default {default_val}): ").strip()
            return int(s) if s else int(default_val)
        except (EOFError, KeyboardInterrupt):
            return int(default_val)
        except Exception:
            return int(default_val)
    max_open_all = _ask_int('Max concurrent open positions (0=unlimited)', max_pos_open_def)
    max_open_longs = _ask_int('Max concurrent open LONGS (0=unlimited)', max_open_longs_def)
    max_open_shorts = _ask_int('Max concurrent open SHORTS (0=unlimited)', max_open_shorts_def)

    return BacktestParams(
        risk_pct=risk, stop_pct=sl, tp_pct=tp,
        fee_pct=fee, slippage_pct=slip,
        max_position_value_pct=max_pos,
        equity_cap=eq_cap,
        time_stop_bars=tstop,
        single_position_mode=single_pos,
        conservative_sizing=conservative,
        max_open_positions=max_open_all,
        max_open_longs=max_open_longs,
        max_open_shorts=max_open_shorts,
    )


def _direction_from_type(t: str) -> int:
    t = (t or "").lower()
    if "hbear" in t:
        return -1
    return 1


def _align_entry_index(dates: pd.DatetimeIndex, signal_ts: pd.Timestamp) -> int:
    # find first index strictly greater than signal_ts
    pos = dates.searchsorted(signal_ts, side='right')
    return int(pos) if pos < len(dates) else -1


def _simulate_trade_long(df: pd.DataFrame, entry_i: int, entry_price: float, stop_pct: float, tp_pct: float):
    stop = entry_price * (1.0 - stop_pct / 100.0)
    tp = entry_price * (1.0 + tp_pct / 100.0)
    max_high = df['high'].iloc[entry_i]
    mae = 0.0
    mfe = 0.0
    bars = 0
    exit_i = None
    exit_price = None
    # iterate forward starting from next bar after entry
    for j in range(entry_i + 1, len(df)):
        bars += 1
        h = df['high'].iloc[j]
        l = df['low'].iloc[j]
        # update trailing
        if h > max_high:
            max_high = h
        trail = max(stop, max_high * (1.0 - stop_pct / 100.0))
        stop = trail
        # conservative: stop first, then tp
        if l <= stop:
            exit_i = j
            exit_price = stop
            break
        if h >= tp:
            exit_i = j
            exit_price = tp
            break
        # running MFE/MAE
        mfe = max(mfe, (h / entry_price) - 1.0)
        mae = min(mae, (l / entry_price) - 1.0)
    if exit_i is None:
        # exit at last close
        exit_i = len(df) - 1
        exit_price = df['close'].iloc[exit_i]
    # finalize mfe/mae at exit bar
    h = df['high'].iloc[exit_i]
    l = df['low'].iloc[exit_i]
    mfe = max(mfe, (h / entry_price) - 1.0)
    mae = min(mae, (l / entry_price) - 1.0)
    return exit_i, exit_price, bars, mfe, mae


def _simulate_trade_short(df: pd.DataFrame, entry_i: int, entry_price: float, stop_pct: float, tp_pct: float):
    stop = entry_price * (1.0 + stop_pct / 100.0)
    tp = entry_price * (1.0 - tp_pct / 100.0)
    min_low = df['low'].iloc[entry_i]
    mae = 0.0
    mfe = 0.0
    bars = 0
    exit_i = None
    exit_price = None
    for j in range(entry_i + 1, len(df)):
        bars += 1
        h = df['high'].iloc[j]
        l = df['low'].iloc[j]
        # update trailing (favorable downwards)
        if l < min_low:
            min_low = l
        trail = min(stop, min_low * (1.0 + stop_pct / 100.0))
        stop = trail
        # conservative: stop first, then tp (short)
        if h >= stop:
            exit_i = j
            exit_price = stop
            break
        if l <= tp:
            exit_i = j
            exit_price = tp
            break
        # running MFE/MAE (for short: invert logic)
        mfe = max(mfe, (entry_price / l) - 1.0)  # favorable down move
        mae = min(mae, 1.0 - (entry_price / h))  # adverse up move negative
    if exit_i is None:
        exit_i = len(df) - 1
        exit_price = df['close'].iloc[exit_i]
    # finalize mfe/mae at exit bar
    h = df['high'].iloc[exit_i]
    l = df['low'].iloc[exit_i]
    mfe = max(mfe, (entry_price / l) - 1.0)
    mae = min(mae, 1.0 - (entry_price / h))
    return exit_i, exit_price, bars, mfe, mae


def _simulate_trade_long_time(df: pd.DataFrame, entry_i: int, entry_price: float, stop_pct: float, tp_pct: float, time_stop: int):
    stop = entry_price * (1.0 - stop_pct / 100.0)
    tp = entry_price * (1.0 + tp_pct / 100.0)
    max_high = df['high'].iloc[entry_i]
    mae = 0.0
    mfe = 0.0
    bars = 0
    exit_i = None
    exit_price = None
    for j in range(entry_i + 1, len(df)):
        bars += 1
        h = df['high'].iloc[j]
        l = df['low'].iloc[j]
        if h > max_high:
            max_high = h
        trail = max(stop, max_high * (1.0 - stop_pct / 100.0))
        stop = trail
        if l <= stop:
            exit_i = j
            exit_price = stop
            break
        if h >= tp:
            exit_i = j
            exit_price = tp
            break
        if bars >= time_stop:
            exit_i = j
            exit_price = df['close'].iloc[j]
            break
        mfe = max(mfe, (h / entry_price) - 1.0)
        mae = min(mae, (l / entry_price) - 1.0)
    if exit_i is None:
        exit_i = len(df) - 1
        exit_price = df['close'].iloc[exit_i]
    h = df['high'].iloc[exit_i]
    l = df['low'].iloc[exit_i]
    mfe = max(mfe, (h / entry_price) - 1.0)
    mae = min(mae, (l / entry_price) - 1.0)
    return exit_i, exit_price, bars, mfe, mae


def _simulate_trade_short_time(df: pd.DataFrame, entry_i: int, entry_price: float, stop_pct: float, tp_pct: float, time_stop: int):
    stop = entry_price * (1.0 + stop_pct / 100.0)
    tp = entry_price * (1.0 - tp_pct / 100.0)
    min_low = df['low'].iloc[entry_i]
    mae = 0.0
    mfe = 0.0
    bars = 0
    exit_i = None
    exit_price = None
    for j in range(entry_i + 1, len(df)):
        bars += 1
        h = df['high'].iloc[j]
        l = df['low'].iloc[j]
        if l < min_low:
            min_low = l
        trail = min(stop, min_low * (1.0 + stop_pct / 100.0))
        stop = trail
        if h >= stop:
            exit_i = j
            exit_price = stop
            break
        if l <= tp:
            exit_i = j
            exit_price = tp
            break
        if bars >= time_stop:
            exit_i = j
            exit_price = df['close'].iloc[j]
            break
        mfe = max(mfe, (entry_price / l) - 1.0)
        mae = min(mae, 1.0 - (entry_price / h))
    if exit_i is None:
        exit_i = len(df) - 1
        exit_price = df['close'].iloc[exit_i]
    h = df['high'].iloc[exit_i]
    l = df['low'].iloc[exit_i]
    mfe = max(mfe, (entry_price / l) - 1.0)
    mae = min(mae, 1.0 - (entry_price / h))
    return exit_i, exit_price, bars, mfe, mae


def backtest(df_ohlc: pd.DataFrame, markers: pd.DataFrame, params: BacktestParams) -> Dict[str, Any]:
    # prepare data
    df = df_ohlc.copy()
    if 'date' not in df.columns:
        raise ValueError("OHLC DataFrame must contain 'date' column")
    df['date'] = pd.to_datetime(df['date'], utc=True, errors='coerce')
    df = df.dropna(subset=['date']).sort_values('date').reset_index(drop=True)
    for col in ['open', 'high', 'low', 'close']:
        if col not in df.columns:
            raise ValueError(f"OHLC DataFrame missing column: {col}")
    dates = pd.DatetimeIndex(df['date'])

    mk = markers.copy()
    if 'Date' not in mk.columns or 'Type' not in mk.columns:
        raise ValueError("Markers DataFrame must contain 'Date' and 'Type'")
    mk['Date'] = pd.to_datetime(mk['Date'], utc=True, errors='coerce')
    mk = mk.dropna(subset=['Date', 'Type']).sort_values('Date').reset_index(drop=True)

    equity = params.starting_capital
    realized_equity = equity
    equity_points = []  # (date, equity) at settlement moments (legacy)
    trades = []
    open_trades = []  # list of dicts with exit_i, pnl, exit_date, dir

    for _, row in mk.iterrows():
        direction = _direction_from_type(row['Type'])
        if direction < 0 and not params.allow_shorts:
            continue
        signal_ts = row['Date']
        ei = _align_entry_index(dates, signal_ts)
        if ei < 0:
            continue
        # settle any trades that exited before this entry index
        if open_trades:
            remaining = []
            for ot in open_trades:
                if ot['exit_i'] < ei:
                    realized_equity += ot['pnl']
                    equity_points.append((ot['exit_date'], realized_equity))
                else:
                    remaining.append(ot)
            open_trades = remaining

        # enforce concurrent caps
        if getattr(params, 'single_position_mode', False) and open_trades:
            continue
        if params.max_open_positions and len(open_trades) >= params.max_open_positions:
            continue
        if params.max_open_longs and sum(1 for ot in open_trades if ot.get('dir') == 'Long') >= params.max_open_longs:
            if direction > 0:
                continue
        if params.max_open_shorts and sum(1 for ot in open_trades if ot.get('dir') == 'Short') >= params.max_open_shorts:
            if direction < 0:
                continue

        entry_price = float(df['open'].iloc[ei])
        # position sizing by risk per trade
        stop_pct = params.stop_pct
        tp_pct = params.tp_pct
        if direction > 0:
            stop_level = entry_price * (1 - stop_pct / 100.0)
            stop_dist = max(entry_price - stop_level, 1e-9)
        else:
            stop_level = entry_price * (1 + stop_pct / 100.0)
            stop_dist = max(stop_level - entry_price, 1e-9)
        # effective equity with optional cap (realized only; no MTM)
        effective_equity = min(realized_equity, params.equity_cap) if params.equity_cap and params.equity_cap > 0 else realized_equity
        risk_cash = effective_equity * (params.risk_pct / 100.0)
        qty = int(max(np.floor(risk_cash / stop_dist), 1))
        # cap position notional as % of equity
        if params.max_position_value_pct and params.max_position_value_pct > 0:
            max_notional = effective_equity * (params.max_position_value_pct / 100.0)
            if entry_price > 0 and max_notional > 0:
                qty_cap = int(max(np.floor(max_notional / entry_price), 1))
                qty = min(qty, qty_cap)
        if qty <= 0:
            continue

        if direction > 0:
            exit_i, exit_price, bars, mfe, mae = _simulate_trade_long(df, ei, entry_price, stop_pct, tp_pct) if params.time_stop_bars <= 0 else _simulate_trade_long_time(df, ei, entry_price, stop_pct, tp_pct, params.time_stop_bars)
        else:
            exit_i, exit_price, bars, mfe, mae = _simulate_trade_short(df, ei, entry_price, stop_pct, tp_pct) if params.time_stop_bars <= 0 else _simulate_trade_short_time(df, ei, entry_price, stop_pct, tp_pct, params.time_stop_bars)

        # Apply slippage adversely, and fees per side on notional
        slip = params.slippage_pct / 100.0
        fee = params.fee_pct / 100.0
        if direction > 0:
            entry_fill = entry_price * (1.0 + slip)
            exit_fill = exit_price * (1.0 - slip)
            gross = (exit_fill - entry_fill) * qty
        else:
            entry_fill = entry_price * (1.0 - slip)
            exit_fill = exit_price * (1.0 + slip)
            gross = (entry_fill - exit_fill) * qty
        fees = fee * (entry_fill * qty + exit_fill * qty)
        pnl = gross - fees

        # Conservative sizing option: recompute qty if using entry_fill for stop distance (more conservative)
        if params.conservative_sizing:
            if direction > 0:
                stop_level = entry_fill * (1 - stop_pct / 100.0)
                stop_dist_fill = max(entry_fill - stop_level, 1e-9)
            else:
                stop_level = entry_fill * (1 + stop_pct / 100.0)
                stop_dist_fill = max(stop_level - entry_fill, 1e-9)
            qty_cons = int(max(np.floor(risk_cash / stop_dist_fill), 1))
            if qty_cons < qty:
                qty = qty_cons
                # recompute PnL/fees based on new qty
                gross = (exit_fill - entry_fill) * qty if direction > 0 else (entry_fill - exit_fill) * qty
                fees = fee * (entry_fill * qty + exit_fill * qty)
                pnl = gross - fees

        exit_date = df['date'].iloc[exit_i]
        # schedule settlement when exit occurs
        open_trades.append({'exit_i': exit_i, 'pnl': pnl, 'exit_date': exit_date, 'dir': 'Long' if direction > 0 else 'Short'})

        # Derive analysis group from Type
        t_lower = str(row['Type']).lower()
        if 'cbulldivg_x2' in t_lower:
            analysis_group = 'CBullDivg_x2'
        elif 'hbeardivg' in t_lower:
            analysis_group = 'HBearDivg'
        elif 'hbulldivg' in t_lower:
            analysis_group = 'HBullDivg'
        else:
            analysis_group = 'CBullDivg'

        # Asset open/close at entry/exit bars for audit
        entry_open = float(df['open'].iloc[ei])
        entry_close_bar = float(df['close'].iloc[ei])
        exit_open_bar = float(df['open'].iloc[exit_i])
        exit_close_bar = float(df['close'].iloc[exit_i])

        # Theoretical max loss estimate at stop (includes slippage and fees)
        if direction > 0:
            stop_fill = entry_price * (1 - stop_pct / 100.0) * (1 - slip)
            worst = (stop_fill - entry_fill) * qty - fee * (entry_fill * qty + stop_fill * qty)
        else:
            stop_fill = entry_price * (1 + stop_pct / 100.0) * (1 + slip)
            worst = (entry_fill - stop_fill) * qty - fee * (entry_fill * qty + stop_fill * qty)

        trades.append({
            'Type': row['Type'],
            'Analysis': analysis_group,
            'Direction': 'Long' if direction > 0 else 'Short',
            'Signal_Date': signal_ts,
            'Entry_Date': df['date'].iloc[ei],
            'Exit_Date': exit_date,
            'Entry': entry_price,
            'Exit': exit_price,
            'Qty': qty,
            'Equity_At_Entry': effective_equity,
            'Risk_Cash': risk_cash,
            'Entry_Open': entry_open,
            'Entry_Close': entry_close_bar,
            'Exit_Open': exit_open_bar,
            'Exit_Close': exit_close_bar,
            'Entry_Fill': entry_fill,
            'Exit_Fill': exit_fill,
            'Gross_PnL_$': gross,
            'Fees_$': fees,
            'PnL_$': pnl,
            'Theoretical_Max_Loss_$': worst,
            'PnL_%': (exit_price / entry_price - 1.0) * (100.0 if direction > 0 else -100.0),
            'Bars': bars,
            'MFE_%': mfe * 100.0 if direction > 0 else mfe * 100.0,
            'MAE_%': mae * 100.0 if direction > 0 else mae * 100.0,
            'Candle_Percent': row.get('Candle_Percent', np.nan),
            'MACD_Percent': row.get('MACD_Percent', np.nan),
            'SourceCSV': row.get('SourceCSV') if 'SourceCSV' in mk.columns else None,
        })

    # settle remaining open trades in chronological order
    if open_trades:
        for ot in sorted(open_trades, key=lambda x: x['exit_i']):
            realized_equity += ot['pnl']
            equity_points.append((ot['exit_date'], realized_equity))

    trades_df = pd.DataFrame(trades)
    if trades_df.empty:
        results = {
            'trades': trades_df,
            'equity': pd.DataFrame(equity_points, columns=['Date', 'Equity']) if equity_points else pd.DataFrame(),
            'summary': pd.DataFrame(),
            'by_type': pd.DataFrame(),
        }
        return results

    # KPIs
    wins = trades_df[trades_df['PnL_$'] > 0]
    losses = trades_df[trades_df['PnL_$'] <= 0]
    total_pnl = trades_df['PnL_$'].sum()
    win_rate = len(wins) / len(trades_df) if len(trades_df) else 0.0
    avg_win = wins['PnL_$'].mean() if len(wins) else 0.0
    avg_loss = losses['PnL_$'].mean() if len(losses) else 0.0
    profit_factor = (wins['PnL_$'].sum() / abs(losses['PnL_$'].sum())) if len(losses) and abs(losses['PnL_$'].sum()) > 1e-9 else np.nan
    expectancy = (win_rate * avg_win) + ((1 - win_rate) * avg_loss)

    # Build full equity curve across all bars using realized PnL at exit dates
    # Aggregate PnL per exit date
    if not trades_df.empty:
        pnl_by_date = trades_df.groupby('Exit_Date')['PnL_$'].sum().to_dict()
        equity_series = []
        cur = params.starting_capital
        for dt, close_price in zip(df['date'], df['close']):
            cur += pnl_by_date.get(dt, 0.0)
            equity_series.append((dt, float(close_price), cur))
        eq_df = pd.DataFrame(equity_series, columns=['Date', 'Price_Close', 'Equity'])
    else:
        eq_df = pd.DataFrame(columns=['Date', 'Price_Close', 'Equity'])

    # simple drawdown
    if not eq_df.empty:
        eq_df['Peak'] = eq_df['Equity'].cummax()
        eq_df['Drawdown'] = (eq_df['Equity'] - eq_df['Peak']) / eq_df['Peak']
        max_dd = eq_df['Drawdown'].min()
    else:
        max_dd = np.nan

    # Use equity curve final value when available
    final_equity = eq_df['Equity'].iloc[-1] if not eq_df.empty else params.starting_capital + total_pnl
    summary = pd.DataFrame([
        {
            'Starting_Capital': params.starting_capital,
            'Final_Equity': final_equity,
            'Total_PnL': total_pnl,
            'Trades': len(trades_df),
            'Win_Rate_%': round(win_rate * 100.0, 2),
            'Avg_Win_$': round(avg_win, 2),
            'Avg_Loss_$': round(avg_loss, 2),
            'Profit_Factor': round(profit_factor, 3) if isinstance(profit_factor, (float, int)) else np.nan,
            'Expectancy_$': round(expectancy, 2),
            'Max_Drawdown_%': round(max_dd * 100.0, 2) if pd.notna(max_dd) else np.nan,
            'Risk_%': params.risk_pct,
            'Stop_%_Trailing': params.stop_pct,
            'Take_Profit_%': params.tp_pct,
            'Fee_%_per_side': params.fee_pct,
            'Slippage_%': params.slippage_pct,
            'Max_Pos_Value_%': params.max_position_value_pct,
            'Equity_Cap': params.equity_cap,
            'Time_Stop_Bars': params.time_stop_bars,
            'Single_Position_Mode': getattr(params, 'single_position_mode', False),
            'Backtest_Start': str(getattr(params, 'backtest_start', '')) if getattr(params, 'backtest_start', None) is not None else '',
            'Backtest_End': str(getattr(params, 'backtest_end', '')) if getattr(params, 'backtest_end', None) is not None else '',
        }
    ])

    # build by_type without named aggregation syntax to allow labels with % and spaces
    g = trades_df.groupby(['Type', 'Direction'])
    by_type = pd.DataFrame({
        'Trades': g['Type'].count(),
        'Win_Rate_%': g['PnL_$'].apply(lambda s: round(100.0 * (s > 0).mean(), 2)),
        'Total_PnL': g['PnL_$'].sum(),
        'Avg_PnL': g['PnL_$'].mean(),
    }).reset_index().sort_values(['Total_PnL'], ascending=False)

    return {
        'trades': trades_df,
        'equity': eq_df[['Date', 'Price_Close', 'Equity']] if not eq_df.empty else eq_df,
        'summary': summary,
        'by_type': by_type,
    }


def export_backtest_xlsx(results: Dict[str, Any], out_path: str):
    os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)
    trades = results.get('trades', pd.DataFrame())
    equity = results.get('equity', pd.DataFrame())
    summary = results.get('summary', pd.DataFrame())
    by_type = results.get('by_type', pd.DataFrame())

    if not _HAS_XLSX:
        # fallback CSVs
        base = os.path.splitext(out_path)[0]
        summary.to_csv(base + '_summary.csv', index=False)
        trades.to_csv(base + '_trades.csv', index=False)
        equity.to_csv(base + '_equity.csv', index=False)
        by_type.to_csv(base + '_bytype.csv', index=False)
        return

    from openpyxl.chart import LineChart, Reference
    try:
        from openpyxl.formatting.rule import ColorScaleRule
        _fmt_ok = True
    except Exception:
        _fmt_ok = False

    # Helper: ensure tz-naive for Excel
    def _tz_naive(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        from pandas.api.types import is_datetime64_any_dtype
        for c in df.columns:
            if is_datetime64_any_dtype(df[c]):
                # convert tz-aware to naive
                try:
                    df[c] = pd.to_datetime(df[c]).dt.tz_localize(None)
                except Exception:
                    try:
                        df[c] = pd.to_datetime(df[c]).dt.tz_convert(None)
                    except Exception:
                        pass
        return df

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        if not summary.empty:
            summary.to_excel(writer, sheet_name='Summary', index=False)
        if not by_type.empty:
            by_type.to_excel(writer, sheet_name='By_Type', index=False)
        if not trades.empty:
            _tz_naive(trades).to_excel(writer, sheet_name='Trades', index=False)
        if not equity.empty:
            _tz_naive(equity).to_excel(writer, sheet_name='EquityCurve', index=False)
            # Add dual-axis chart: Price (primary) and Equity (secondary)
            ws = writer.sheets['EquityCurve']
            if len(equity) >= 2:
                # X-axis categories = Dates
                cats = Reference(ws, min_col=1, min_row=2, max_row=len(equity) + 1)
                # Price series (primary)
                price_ref = Reference(ws, min_col=2, min_row=1, max_row=len(equity) + 1)
                c1 = LineChart()
                c1.title = 'Price and Equity vs Time'
                c1.x_axis.title = 'Date'
                c1.y_axis.title = 'Price'
                c1.add_data(price_ref, titles_from_data=True)
                c1.set_categories(cats)
                # Equity series on secondary axis
                eq_ref = Reference(ws, min_col=3, min_row=1, max_row=len(equity) + 1)
                c2 = LineChart()
                c2.y_axis.title = 'Equity ($)'
                c2.add_data(eq_ref, titles_from_data=True)
                # set secondary axis by combining charts
                c2.y_axis.axId = 200
                c1.y_axis.axId = 100
                c1 += c2
                # ensure legend is visible (default is True)
                if not getattr(c1, 'legend', None):
                    from openpyxl.chart.legend import Legend
                    c1.legend = Legend()
                ws.add_chart(c1, 'E2')

        # DOE performance pivots (if Candle_Percent & MACD_Percent available)
        trades = results.get('trades', pd.DataFrame())
        if not trades.empty and {'Candle_Percent', 'MACD_Percent'}.issubset(set(trades.columns)):
            sheet = 'DOE_Perf'
            # Aggregate per DOE cell
            grp = trades.groupby(['Candle_Percent', 'MACD_Percent'])
            perf = pd.DataFrame({
                'Trades': grp.size(),
                'Win_Rate_%': grp['PnL_$'].apply(lambda s: round(100.0 * (s > 0).mean(), 2)),
                'Total_PnL': grp['PnL_$'].sum(),
                'Profit_Factor': grp['PnL_$'].apply(lambda s: (s[s>0].sum() / abs(s[s<=0].sum())) if (s[s<=0].sum()) != 0 else np.nan),
            }).reset_index()

            # Write pivots stacked with titles
            start_row = 0
            def write_pivot(title, column):
                nonlocal start_row
                if perf.empty:
                    return
                pivot = perf.pivot(index='Candle_Percent', columns='MACD_Percent', values=column).sort_index().sort_index(axis=1)
                if pivot.empty:
                    return
                pd.DataFrame({title: []}).to_excel(writer, sheet_name=sheet, startrow=start_row, index=False)
                pivot.to_excel(writer, sheet_name=sheet, startrow=start_row + 1)
                if _fmt_ok:
                    ws = writer.sheets[sheet]
                    rows_n, cols_n = pivot.shape
                    if rows_n and cols_n:
                        from openpyxl.utils import get_column_letter
                        r1, c1 = start_row + 2, 2
                        r2, c2 = start_row + 1 + rows_n, 1 + cols_n
                        ref = f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
                        ws.conditional_formatting.add(ref, ColorScaleRule(start_type='min', start_color='93c47d',
                                                                          mid_type='percentile', mid_value=50, mid_color='ffd966',
                                                                          end_type='max', end_color='e06666'))
                        ws.freeze_panes = ws['B2']
                start_row += pivot.shape[0] + 4

            write_pivot('DOE Trades count', 'Trades')
            write_pivot('DOE Win Rate %', 'Win_Rate_%')
            write_pivot('DOE Total PnL', 'Total_PnL')
            write_pivot('DOE Profit Factor', 'Profit_Factor')

            # Per-file performance overview (when multiple DOE marker files are combined)
            if 'SourceCSV' in trades.columns:
                file_grp = trades.groupby('SourceCSV')
                file_perf = pd.DataFrame({
                    'Trades': file_grp.size(),
                    'Win_Rate_%': file_grp['PnL_$'].apply(lambda s: round(100.0 * (s > 0).mean(), 2)),
                    'Total_PnL': file_grp['PnL_$'].sum(),
                    'Profit_Factor': file_grp['PnL_$'].apply(lambda s: (s[s>0].sum() / abs(s[s<=0].sum())) if (s[s<=0].sum()) != 0 else np.nan),
                }).reset_index().sort_values('Total_PnL', ascending=False)
                file_perf.to_excel(writer, sheet_name='DOE_File_Perf', index=False)

            # Per-analysis DOE performance sheets
            analyses = ['CBullDivg', 'CBullDivg_x2', 'HBullDivg', 'HBearDivg']
            for a in analyses:
                sub = trades[trades.get('Analysis') == a]
                if sub.empty:
                    continue
                sub_grp = sub.groupby(['Candle_Percent', 'MACD_Percent'])
                sub_perf = pd.DataFrame({
                    'Trades': sub_grp.size(),
                    'Win_Rate_%': sub_grp['PnL_$'].apply(lambda s: round(100.0 * (s > 0).mean(), 2)),
                    'Total_PnL': sub_grp['PnL_$'].sum(),
                    'Profit_Factor': sub_grp['PnL_$'].apply(lambda s: (s[s>0].sum() / abs(s[s<=0].sum())) if (s[s<=0].sum()) != 0 else np.nan),
                }).reset_index()

                start_row = 0
                sub_sheet = f'DOE_Perf_{a}'
                def write_sub_pivot(title, column):
                    nonlocal start_row
                    pivot = sub_perf.pivot(index='Candle_Percent', columns='MACD_Percent', values=column).sort_index().sort_index(axis=1)
                    if pivot.empty:
                        return
                    pd.DataFrame({f"{a} – {title}": []}).to_excel(writer, sheet_name=sub_sheet, startrow=start_row, index=False)
                    pivot.to_excel(writer, sheet_name=sub_sheet, startrow=start_row + 1)
                    if _fmt_ok:
                        ws = writer.sheets[sub_sheet]
                        rows_n, cols_n = pivot.shape
                        if rows_n and cols_n:
                            from openpyxl.utils import get_column_letter
                            r1, c1 = start_row + 2, 2
                            r2, c2 = start_row + 1 + rows_n, 1 + cols_n
                            ref = f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
                            ws.conditional_formatting.add(ref, ColorScaleRule(start_type='min', start_color='93c47d',
                                                                              mid_type='percentile', mid_value=50, mid_color='ffd966',
                                                                              end_type='max', end_color='e06666'))
                            ws.freeze_panes = ws['B2']
                    start_row += pivot.shape[0] + 4

                write_sub_pivot('Trades count', 'Trades')
                write_sub_pivot('Win Rate %', 'Win_Rate_%')
                write_sub_pivot('Total PnL', 'Total_PnL')
                write_sub_pivot('Profit Factor', 'Profit_Factor')
